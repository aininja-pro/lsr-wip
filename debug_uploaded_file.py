#!/usr/bin/env python3
import sys
sys.path.append('/app/src')
from data_processing.excel_integration_v2 import find_or_create_monthly_tab, load_wip_workbook, find_section_markers
import logging
import openpyxl

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')

print("=== DEBUGGING UPLOADED FILE PROCESSING ===")

# Simulate what the UI does when processing uploaded files
def debug_file_processing(file_path, month_year):
    print(f"\n1. Loading workbook from: {file_path}")
    wb = load_wip_workbook(file_path)
    print(f"   Available sheets: {wb.sheetnames}")
    
    print(f"\n2. Finding/creating monthly tab for: '{month_year}'")
    ws = find_or_create_monthly_tab(wb, month_year)
    print(f"   Selected tab: '{ws.title}'")
    print(f"   Tab dimensions: {ws.max_row} rows x {ws.max_column} columns")
    
    print(f"\n3. Checking specific cells in tab '{ws.title}':")
    # Check where we expect the headers to be
    for row in [1, 2, 3, 69, 70, 71]:
        for col in [1, 2, 3]:
            cell = ws.cell(row=row, column=col)
            cell_value = str(cell.value) if cell.value else "None"
            print(f"   Row {row}, Col {col}: '{cell_value}'")
    
    print(f"\n4. Testing section detection:")
    sections = find_section_markers(ws, ['5040', '5030'])
    print(f"   Results: {sections}")
    
    print(f"\n5. Manual search for '5040' anywhere in the sheet:")
    found_5040 = False
    for row in range(1, min(ws.max_row + 1, 100)):
        for col in range(1, min(ws.max_column + 1, 10)):
            cell = ws.cell(row=row, column=col)
            if cell.value and '5040' in str(cell.value):
                print(f"   Found '5040' at row {row}, col {col}: '{cell.value}'")
                found_5040 = True
    
    if not found_5040:
        print("   No '5040' found in first 100 rows, 10 columns!")
    
    print(f"\n6. Manual search for '5030' anywhere in the sheet:")
    found_5030 = False
    for row in range(1, min(ws.max_row + 1, 100)):
        for col in range(1, min(ws.max_column + 1, 10)):
            cell = ws.cell(row=row, column=col)
            if cell.value and '5030' in str(cell.value):
                print(f"   Found '5030' at row {row}, col {col}: '{cell.value}'")
                found_5030 = True
    
    if not found_5030:
        print("   No '5030' found in first 100 rows, 10 columns!")

# Test with different month formats
test_cases = [
    ("Apr 25", "/app/test_data/Master WIP Report.xlsx"),
    ("April 2025", "/app/test_data/Master WIP Report.xlsx")
]

for month_format, file_path in test_cases:
    print(f"\n{'='*60}")
    print(f"TESTING: Month '{month_format}' with file {file_path}")
    print(f"{'='*60}")
    try:
        debug_file_processing(file_path, month_format)
    except Exception as e:
        print(f"ERROR: {e}")

print(f"\n{'='*60}")
print("DEBUGGING COMPLETE")
print("Compare the results above to see what's different!")
print(f"{'='*60}") 