#!/usr/bin/env python3
"""
Test the entire file processing pipeline to find corruption point
"""
import sys
from pathlib import Path
import os
import io
import pandas as pd
sys.path.append(str(Path(__file__).parent))

from data_processing.excel_integration_v2 import (
    load_wip_workbook, 
    find_or_create_monthly_tab,
    update_wip_report_v2
)
import openpyxl
import logging

# Set up logging
logging.basicConfig(level=logging.DEBUG, format='%(levelname)s:%(name)s:%(message)s')
logger = logging.getLogger(__name__)

def test_full_pipeline():
    """Test the complete file processing pipeline"""
    print("🧪 Testing Full WIP Processing Pipeline...")
    
    # Test file paths
    original_file = "/app/test_data/Master WIP Report.xlsx"
    test_file = "/app/test_master_pipeline.xlsx"
    
    try:
        # Step 1: Copy original file for testing
        print("\n1️⃣ Copying original file for testing...")
        with open(original_file, "rb") as src, open(test_file, "wb") as dst:
            dst.write(src.read())
        print(f"   ✅ Test file created: {test_file}")
        
        # Step 2: Verify original file integrity
        print("\n2️⃣ Verifying original file integrity...")
        try:
            wb_original = openpyxl.load_workbook(test_file, keep_vba=True)
            print(f"   ✅ Original file loads successfully")
            print(f"   Sheets: {len(wb_original.sheetnames)} found")
            wb_original.close()
        except Exception as e:
            print(f"   ❌ Original file failed to load: {e}")
            return
        
        # Step 3: Create test data for update
        print("\n3️⃣ Creating test data...")
        sub_labor_data = pd.DataFrame({
            'Job Number': ['TEST001', 'TEST002'],
            'Sub Labor': [1000.0, 2000.0]
        })
        
        material_data = pd.DataFrame({
            'Job Number': ['TEST001', 'TEST002'], 
            'Material': [500.0, 750.0]
        })
        print(f"   ✅ Test data created: {len(sub_labor_data)} sub labor, {len(material_data)} material")
        
        # Step 4: Test the update_wip_report_v2 function
        print("\n4️⃣ Testing update_wip_report_v2 function...")
        try:
            result = update_wip_report_v2(
                file_path=test_file,
                sub_labor_data=sub_labor_data,
                material_data=material_data,
                month_year="Apr 25",
                create_backup_flag=False  # Skip backup for testing
            )
            
            if result['success']:
                print(f"   ✅ Update function completed successfully")
                if result.get('sections_found'):
                    print(f"   ✅ Sections found: {result['sections_found']}")
                else:
                    print(f"   ⚠️ No sections info returned")
            else:
                print(f"   ❌ Update function failed: {result.get('error', 'Unknown error')}")
                return
                
        except Exception as e:
            print(f"   ❌ Update function crashed: {e}")
            import traceback
            traceback.print_exc()
            return
        
        # Step 5: Test if updated file is valid
        print("\n5️⃣ Testing updated file integrity...")
        try:
            wb_updated = openpyxl.load_workbook(test_file, keep_vba=True)
            print(f"   ✅ Updated file loads successfully")
            print(f"   Sheets: {len(wb_updated.sheetnames)} found")
            
            # Check if Apr 25 sheet exists and has data
            if "Apr 25" in wb_updated.sheetnames:
                ws = wb_updated["Apr 25"]
                print(f"   ✅ 'Apr 25' sheet exists with {ws.max_row} rows, {ws.max_column} columns")
            else:
                print(f"   ⚠️ 'Apr 25' sheet not found")
            
            wb_updated.close()
        except Exception as e:
            print(f"   ❌ Updated file failed to load: {e}")
            return
        
        # Step 6: Test file read for download (simulating Streamlit)
        print("\n6️⃣ Testing download simulation...")
        try:
            # Method 1: Raw file read (old way)
            with open(test_file, "rb") as f:
                raw_bytes = f.read()
            print(f"   📄 Raw file size: {len(raw_bytes)} bytes")
            
            # Method 2: BytesIO buffer (new way)
            excel_buffer = io.BytesIO(raw_bytes)
            excel_buffer.seek(0)
            buffered_bytes = excel_buffer.getvalue()
            excel_buffer.close()
            print(f"   💾 Buffered size: {len(buffered_bytes)} bytes")
            
            # Test if buffered bytes work
            test_buffer = io.BytesIO(buffered_bytes)
            wb_test = openpyxl.load_workbook(test_buffer, keep_vba=True)
            print(f"   ✅ Buffered bytes work for openpyxl")
            wb_test.close()
            test_buffer.close()
            
        except Exception as e:
            print(f"   ❌ Download simulation failed: {e}")
            import traceback
            traceback.print_exc()
            return
        
        # Step 7: Save test download file
        print("\n7️⃣ Creating test download file...")
        download_file = "/app/test_download_result.xlsx"
        try:
            with open(download_file, "wb") as f:
                f.write(buffered_bytes)
            print(f"   💾 Test download saved: {download_file}")
            
            # Verify test download file
            wb_download = openpyxl.load_workbook(download_file, keep_vba=True)
            print(f"   ✅ Test download file works!")
            wb_download.close()
            
        except Exception as e:
            print(f"   ❌ Test download file failed: {e}")
            return
        
        print("\n🎉 ALL TESTS PASSED! The pipeline works correctly.")
        print("\nFiles for manual testing:")
        print(f"   - Original: {original_file}")
        print(f"   - Updated: {test_file}")
        print(f"   - Download: {download_file}")
        
    except Exception as e:
        print(f"❌ Pipeline test failed: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        # Cleanup
        for temp_file in [test_file, "/app/test_download_result.xlsx"]:
            if os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                    print(f"   🧹 Cleaned up: {temp_file}")
                except:
                    pass

if __name__ == "__main__":
    test_full_pipeline() 