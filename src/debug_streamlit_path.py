#!/usr/bin/env python3
"""
Debug Streamlit file paths and working directory
"""
import os
import sys
from pathlib import Path
import io
import openpyxl

def debug_streamlit_environment():
    """Debug the Streamlit environment"""
    print("🔍 Debugging Streamlit Environment...")
    
    print(f"\n📁 Working Directory: {os.getcwd()}")
    print(f"📁 __file__ location: {__file__}")
    print(f"📁 sys.path: {sys.path[:3]}...")  # First 3 entries
    
    # Test file paths
    test_file = "/app/test_data/Master WIP Report.xlsx"
    relative_temp = "temp_master_debug.xlsx"
    absolute_temp = "/app/temp_master_debug.xlsx"
    
    print(f"\n📄 Test file exists: {os.path.exists(test_file)}")
    
    # Test creating files in different locations
    try:
        # Test 1: Relative path (what Streamlit app uses)
        print(f"\n🧪 Testing relative path: {relative_temp}")
        with open(test_file, "rb") as src:
            file_data = src.read()
        
        with open(relative_temp, "wb") as dst:
            dst.write(file_data)
        
        print(f"   ✅ Created: {os.path.abspath(relative_temp)}")
        print(f"   📏 Size: {os.path.getsize(relative_temp)} bytes")
        
        # Test if it can be loaded
        wb = openpyxl.load_workbook(relative_temp, keep_vba=True)
        print(f"   ✅ Loads with openpyxl: {len(wb.sheetnames)} sheets")
        wb.close()
        
        # Test the download process
        with open(relative_temp, "rb") as f:
            read_data = f.read()
        
        # BytesIO process (what our fix does)
        excel_buffer = io.BytesIO(read_data)
        excel_buffer.seek(0)
        buffered_bytes = excel_buffer.getvalue()
        excel_buffer.close()
        
        print(f"   📏 Read size: {len(read_data)} bytes")
        print(f"   📏 Buffer size: {len(buffered_bytes)} bytes")
        
        # Test final download file
        download_test = "download_test.xlsx"
        with open(download_test, "wb") as f:
            f.write(buffered_bytes)
        
        wb_test = openpyxl.load_workbook(download_test, keep_vba=True)
        print(f"   ✅ Download simulation works: {len(wb_test.sheetnames)} sheets")
        wb_test.close()
        
        # Cleanup
        os.remove(relative_temp)
        os.remove(download_test)
        print(f"   🧹 Cleaned up test files")
        
    except Exception as e:
        print(f"   ❌ Error: {e}")
        import traceback
        traceback.print_exc()
    
    # Test 2: Absolute path
    try:
        print(f"\n🧪 Testing absolute path: {absolute_temp}")
        with open(test_file, "rb") as src, open(absolute_temp, "wb") as dst:
            dst.write(src.read())
        
        print(f"   ✅ Created: {absolute_temp}")
        print(f"   📏 Size: {os.path.getsize(absolute_temp)} bytes")
        
        wb = openpyxl.load_workbook(absolute_temp, keep_vba=True)
        print(f"   ✅ Loads with openpyxl: {len(wb.sheetnames)} sheets")
        wb.close()
        
        os.remove(absolute_temp)
        print(f"   🧹 Cleaned up")
        
    except Exception as e:
        print(f"   ❌ Error: {e}")
    
    print("\n✅ Environment debugging complete!")

if __name__ == "__main__":
    debug_streamlit_environment() 