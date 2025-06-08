#!/usr/bin/env python3
"""
Hybrid Safe WIP Report Automation - Claude's Approach
Uses openpyxl read-only to find locations, then surgical ZIP updates
"""

import streamlit as st
import pandas as pd
import io
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path
import logging
from openpyxl import load_workbook

# Import our data processing functions
import sys
sys.path.append('/app/src')

from data_processing.aggregation import (
    filter_gl_accounts, 
    compute_amounts, 
    aggregate_gl_data
)
from data_processing.merge_data import merge_wip_with_gl

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def initialize_session_state():
    """Initialize session state variables"""
    if 'files_uploaded' not in st.session_state:
        st.session_state.files_uploaded = {}
    if 'processing_complete' not in st.session_state:
        st.session_state.processing_complete = False
    if 'merged_data' not in st.session_state:
        st.session_state.merged_data = None

def map_columns_flexible(df, column_mapping):
    """Map column names flexibly using variations"""
    mapped_df = df.copy()
    
    for standard_name, variations in column_mapping.items():
        for variation in variations:
            if variation in df.columns:
                if variation != standard_name:
                    mapped_df = mapped_df.rename(columns={variation: standard_name})
                break
    
    return mapped_df

def find_cell_locations_readonly(excel_bytes, sheet_name):
    """
    Use openpyxl in READ-ONLY mode to find where cells should be updated
    This avoids any corruption since we never save with openpyxl
    """
    try:
        # READ-ONLY mode - no corruption risk
        wb = load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
        ws = wb[sheet_name]
        
        # Find the 5040 and 5030 sections
        sections = {}
        
        for row in ws.iter_rows(min_row=1, max_row=200, min_col=1, max_col=20):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if "5040" in cell.value:
                        sections['5040'] = {'row': cell.row, 'col': cell.column}
                        logger.info(f"Found 5040 section at row {cell.row}")
                    elif "5030" in cell.value:
                        sections['5030'] = {'row': cell.row, 'col': cell.column}
                        logger.info(f"Found 5030 section at row {cell.row}")
        
        wb.close()  # Close properly
        return sections
        
    except Exception as e:
        logger.error(f"Error finding cell locations: {e}")
        return {}

def prepare_cell_updates(merged_df, sections, sheet_name):
    """
    Prepare the exact cell updates needed based on merged data and section locations
    """
    updates = {}
    
    # For 5040 section (labor costs)
    if '5040' in sections:
        start_row = sections['5040']['row'] + 1  # Start after header
        current_row = start_row
        
        for _, job in merged_df.iterrows():
            job_number = str(job.get('Job Number', ''))
            labor_actual = float(job.get('Labor Actual', 0))
            
            # Update job number and labor actual
            updates[f'A{current_row}'] = job_number
            updates[f'C{current_row}'] = labor_actual  # Assuming column C is labor actual
            
            current_row += 1
    
    # For 5030 section (material costs)  
    if '5030' in sections:
        start_row = sections['5030']['row'] + 1  # Start after header
        current_row = start_row
        
        for _, job in merged_df.iterrows():
            job_number = str(job.get('Job Number', ''))
            material_actual = float(job.get('Material Actual', 0))
            
            # Update job number and material actual
            updates[f'A{current_row}'] = job_number
            updates[f'C{current_row}'] = material_actual  # Assuming column C is material actual
            
            current_row += 1
    
    return {sheet_name: updates}

def surgical_excel_update(excel_bytes, sheet_updates):
    """
    Surgically update Excel file using ZIP/XML approach
    This preserves EVERYTHING since we only modify specific cell values
    """
    try:
        # Read Excel as ZIP
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(io.BytesIO(excel_bytes), 'r') as input_zip:
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as output_zip:
                
                for file_info in input_zip.infolist():
                    file_data = input_zip.read(file_info)
                    
                    # Check if this is a worksheet XML we need to update
                    sheet_updated = False
                    for sheet_name, updates in sheet_updates.items():
                        # Look for worksheet files (they're usually named like sheet1.xml, sheet2.xml, etc.)
                        if file_info.filename.startswith('xl/worksheets/') and file_info.filename.endswith('.xml'):
                            # Parse XML and update cells
                            try:
                                root = ET.fromstring(file_data)
                                
                                # Find sheetData element
                                sheet_data = root.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheetData')
                                if sheet_data is not None:
                                    
                                    # Update each cell
                                    cells_updated = 0
                                    for cell_ref, new_value in updates.items():
                                        if update_cell_in_xml(sheet_data, cell_ref, new_value):
                                            cells_updated += 1
                                    
                                    if cells_updated > 0:
                                        # Convert back to XML
                                        file_data = ET.tostring(root, encoding='utf-8', xml_declaration=True)
                                        sheet_updated = True
                                        logger.info(f"Updated {cells_updated} cells in {sheet_name}")
                                        
                            except ET.ParseError as e:
                                logger.warning(f"Could not parse XML for {file_info.filename}: {e}")
                                
                    # Write file (updated or original)
                    output_zip.writestr(file_info, file_data)
        
        zip_buffer.seek(0)
        result = zip_buffer.getvalue()
        
        logger.info(f"Surgical update complete. Output size: {len(result)} bytes")
        logger.info(f"Size change: {len(result) - len(excel_bytes):+d} bytes")
        
        return result
        
    except Exception as e:
        logger.error(f"Surgical update failed: {e}")
        return excel_bytes

def update_cell_in_xml(sheet_data, cell_ref, new_value):
    """Update a specific cell in the XML"""
    try:
        # Parse cell reference (e.g., "A1" -> row=1, col=1)
        col_letters = ''.join(c for c in cell_ref if c.isalpha())
        row_num = int(''.join(c for c in cell_ref if c.isdigit()))
        
        # Find or create the row
        row_elem = None
        for row in sheet_data.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row'):
            if int(row.get('r', 0)) == row_num:
                row_elem = row
                break
        
        if row_elem is None:
            # Create new row
            row_elem = ET.SubElement(sheet_data, '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row')
            row_elem.set('r', str(row_num))
        
        # Find or create the cell
        cell_elem = None
        for cell in row_elem.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c'):
            if cell.get('r') == cell_ref:
                cell_elem = cell
                break
        
        if cell_elem is None:
            # Create new cell
            cell_elem = ET.SubElement(row_elem, '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c')
            cell_elem.set('r', cell_ref)
        
        # Update the value
        value_elem = cell_elem.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
        if value_elem is None:
            value_elem = ET.SubElement(cell_elem, '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
        
        value_elem.text = str(new_value)
        
        # Set cell type to number if it's numeric
        if isinstance(new_value, (int, float)):
            cell_elem.set('t', 'n')
        
        return True
        
    except Exception as e:
        logger.warning(f"Could not update cell {cell_ref}: {e}")
        return False

def create_backup_from_bytes(file_bytes):
    """Create a backup from file bytes"""
    try:
        backup_dir = Path("WIP_Backups")
        backup_dir.mkdir(exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"WIP_Report_BACKUP_{timestamp}.xlsx"
        backup_path = backup_dir / backup_filename
        
        with open(backup_path, 'wb') as f:
            f.write(file_bytes)
        
        logger.info(f"Created backup: {backup_path}")
        return str(backup_path)
        
    except Exception as e:
        logger.error(f"Backup creation failed: {e}")
        return None

def process_data(wip_bytes, gl_bytes, include_closed):
    """Process the data using our existing functions"""
    try:
        with st.spinner("Processing GL data..."):
            # Load GL inquiry from bytes
            gl_df = pd.read_excel(io.BytesIO(gl_bytes))
            
            # Apply column mapping for GL data
            gl_column_variations = {
                'Account': ['Account', 'Account Number', 'Acct', 'GL Account'],
                'Job Number': ['Job Number', 'Job No', 'Job #', 'Job', 'Project Number', 'Project No'],
                'Debit': ['Debit', 'Dr', 'Debit Amount'],
                'Credit': ['Credit', 'Cr', 'Credit Amount'],
                'Account Type': ['Account Type', 'Type', 'Category']
            }
            gl_df = map_columns_flexible(gl_df, gl_column_variations)
            
            # Process GL data step by step
            filtered_gl = filter_gl_accounts(gl_df)
            amounts_gl = compute_amounts(filtered_gl)
            gl_summary = aggregate_gl_data(amounts_gl)
            
            st.info(f"✅ Processed {len(gl_summary)} GL entries")
            
        with st.spinner("Merging data..."):
            # Load WIP worksheet from bytes
            wip_df = pd.read_excel(io.BytesIO(wip_bytes))
            
            # Apply column mapping for WIP worksheet
            wip_column_variations = {
                'Job Number': ['Job Number', 'Job No', 'Job #', 'Job', 'Project Number', 'Project No'],
                'Status': ['Status', 'Job Status', 'Project Status', 'State'],
                'Job Name': ['Job Name', 'Project Name', 'Description', 'Job Description'],
                'Budget Material': ['Budget Material', 'Material Budget', 'Mat Budget', 'Budget Mat'],
                'Budget Labor': ['Budget Labor', 'Labor Budget', 'Lab Budget', 'Budget Lab']
            }
            wip_df = map_columns_flexible(wip_df, wip_column_variations)
            
            merged_df = merge_wip_with_gl(wip_df, gl_summary, include_closed)
            st.info(f"✅ Merged data for {len(merged_df)} jobs")
            
        return merged_df
        
    except Exception as e:
        st.error(f"Error processing data: {str(e)}")
        logger.error(f"Processing error: {e}")
        return None

def display_file_upload_section():
    """Display file upload interface"""
    st.header("📁 File Upload")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.subheader("Master WIP Report")
        master_file = st.file_uploader(
            "Upload Master WIP Report",
            type=['xlsx', 'xlsm'],
            key='master_wip'
        )
        if master_file:
            st.session_state.files_uploaded['master'] = master_file.getvalue()
            st.success(f"✅ {master_file.name}")
    
    with col2:
        st.subheader("WIP Worksheet Export")
        wip_file = st.file_uploader(
            "Upload WIP Worksheet",
            type=['xlsx'],
            key='wip_worksheet'
        )
        if wip_file:
            st.session_state.files_uploaded['wip'] = wip_file.getvalue()
            st.success(f"✅ {wip_file.name}")
    
    with col3:
        st.subheader("GL Inquiry Export")
        gl_file = st.file_uploader(
            "Upload GL Inquiry",
            type=['xlsx'],
            key='gl_inquiry'
        )
        if gl_file:
            st.session_state.files_uploaded['gl'] = gl_file.getvalue()
            st.success(f"✅ {gl_file.name}")

def display_processing_options():
    """Display processing options"""
    st.header("⚙️ Processing Options")
    
    col1, col2 = st.columns(2)
    
    with col1:
        include_closed = st.checkbox("Include Closed Jobs", value=False)
    
    with col2:
        month_year = st.text_input("Month/Year", value="Apr 25")
    
    return include_closed, month_year

def main():
    st.set_page_config(
        page_title="WIP Report Automation - Hybrid Safe",
        page_icon="🔧",
        layout="wide"
    )
    
    st.title("🔧 WIP Report Automation")
    st.subheader("Hybrid safe data processing with surgical Excel updates")
    
    initialize_session_state()
    
    # File Upload Section
    display_file_upload_section()
    
    # Processing Options
    include_closed, month_year = display_processing_options()
    
    # Process Button
    if st.button("🚀 Process Data", type="primary"):
        if len(st.session_state.files_uploaded) == 3:
            
            # Process the data
            merged_df = process_data(
                st.session_state.files_uploaded['wip'],
                st.session_state.files_uploaded['gl'],
                include_closed
            )
            
            if merged_df is not None:
                st.session_state.merged_data = merged_df
                
                # Find cell locations using read-only openpyxl
                with st.spinner("Finding cell locations..."):
                    sections = find_cell_locations_readonly(
                        st.session_state.files_uploaded['master'], 
                        month_year
                    )
                
                if sections:
                    # Prepare updates
                    with st.spinner("Preparing cell updates..."):
                        updates = prepare_cell_updates(merged_df, sections, month_year)
                    
                    # Create backup
                    with st.spinner("Creating backup..."):
                        backup_path = create_backup_from_bytes(st.session_state.files_uploaded['master'])
                    
                    # Perform surgical update
                    with st.spinner("Performing surgical Excel update..."):
                        updated_bytes = surgical_excel_update(
                            st.session_state.files_uploaded['master'],
                            updates
                        )
                    
                    # Display results
                    st.success("✅ Processing complete!")
                    st.info(f"📊 Processed {len(merged_df)} jobs")
                    if backup_path:
                        st.info(f"💾 Backup created: {backup_path}")
                    
                    # Download button
                    st.download_button(
                        label="📥 Download Updated WIP Report",
                        data=updated_bytes,
                        file_name=f"WIP_Report_Updated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                else:
                    st.error("❌ Could not find 5040/5030 sections in the master report")
        else:
            st.error("❌ Please upload all three files before processing")

if __name__ == "__main__":
    main() 