#!/usr/bin/env python3
"""
Safe WIP Report Automation - Report Generation Only
Generates update reports that can be manually copied into Excel
This avoids ALL Excel corruption issues
"""

import streamlit as st
import pandas as pd
import io
from datetime import datetime
from pathlib import Path
import logging

# Import our data processing functions
import sys
sys.path.append('/app/src')

from data_processing.aggregation import (
    filter_gl_accounts, 
    compute_amounts, 
    aggregate_gl_data
)
from data_processing.merge_data import merge_wip_with_gl

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def initialize_session_state():
    """Initialize session state variables"""
    if 'files_uploaded' not in st.session_state:
        st.session_state.files_uploaded = {}
    if 'processing_complete' not in st.session_state:
        st.session_state.processing_complete = False
    if 'merged_data' not in st.session_state:
        st.session_state.merged_data = None
    if 'results_ready' not in st.session_state:
        st.session_state.results_ready = False
    if 'labor_df' not in st.session_state:
        st.session_state.labor_df = None
    if 'material_df' not in st.session_state:
        st.session_state.material_df = None
    if 'excel_report' not in st.session_state:
        st.session_state.excel_report = None
    if 'month_year' not in st.session_state:
        st.session_state.month_year = None
    if 'gl_entries' not in st.session_state:
        st.session_state.gl_entries = 0

def map_columns_flexible(df, column_mapping):
    """Map column names flexibly using variations"""
    mapped_df = df.copy()
    
    for standard_name, variations in column_mapping.items():
        for variation in variations:
            if variation in df.columns:
                if variation != standard_name:
                    mapped_df = mapped_df.rename(columns={variation: standard_name})
                break
    
    return mapped_df

def process_data(wip_bytes, gl_bytes, include_closed):
    """Process the data using our existing functions"""
    try:
        with st.spinner("Processing GL data..."):
            # Load GL inquiry from bytes
            gl_df = pd.read_excel(io.BytesIO(gl_bytes))
            
            # Log available GL columns to help debug
            logger.info(f"Available GL Inquiry columns: {list(gl_df.columns)}")
            
            # Apply column mapping for GL data
            gl_column_variations = {
                'Account': ['Account', 'Account Number', 'Acct', 'GL Account'],
                'Job Number': ['Job Number', 'Job No', 'Job #', 'Job', 'Project Number', 'Project No'],
                'Debit': ['Debit', 'Dr', 'Debit Amount'],
                'Credit': ['Credit', 'Cr', 'Credit Amount'],
                'Account Type': ['Account Type', 'Type', 'Category']
            }
            gl_df = map_columns_flexible(gl_df, gl_column_variations)
            
            # Process GL data step by step
            filtered_gl = filter_gl_accounts(gl_df)
            amounts_gl = compute_amounts(filtered_gl)
            gl_summary = aggregate_gl_data(amounts_gl)
            
            # Store GL entries count for results display
            st.session_state.gl_entries = len(gl_summary)
            
        with st.spinner("Merging data..."):
            # Load WIP worksheet from bytes
            wip_df = pd.read_excel(io.BytesIO(wip_bytes))
            
            # Log available columns to help debug
            logger.info(f"Available WIP Worksheet columns: {list(wip_df.columns)}")
            
            # Apply column mapping for WIP worksheet
            wip_column_variations = {
                'Job Number': ['Job Number', 'Job No', 'Job #', 'Job', 'Project Number', 'Project No'],
                'Status': ['Status', 'Job Status', 'Project Status', 'State'],
                'Job Name': ['Job Name', 'Project Name', 'Description', 'Job Description'],
                'Budget Material': ['Budget Material', 'Material Budget', 'Mat Budget', 'Budget Mat'],
                'Budget Labor': ['Budget Labor', 'Labor Budget', 'Lab Budget', 'Budget Lab'],
                'Contract Amount': ['Contract Amount', 'Contract Value', 'Total Contract', 'Contract'],
                'Estimated Sub Labor': ['Estimated Sub Labor', 'Est Sub Labor', 'Sub Labor Budget', 'Sub Labor Est'],
                'Estimated Material': ['Estimated Material', 'Est Material', 'Material Budget', 'Material Est']
            }
            wip_df = map_columns_flexible(wip_df, wip_column_variations)
            
            # Log mapped columns
            logger.info(f"WIP Worksheet columns after mapping: {list(wip_df.columns)}")
            
            merged_df = merge_wip_with_gl(wip_df, gl_summary, include_closed)
            
        return merged_df
        
    except Exception as e:
        st.error(f"Error processing data: {str(e)}")
        logger.error(f"Processing error: {e}")
        return None

def generate_update_reports(merged_df):
    """Generate reports with EXACTLY the fields requested"""
    
    # 5040 Section - Labor Report (with Percent Complete column)
    labor_data = []
    for _, job in merged_df.iterrows():
        labor_actual = job.get('5040', 0) or job.get('Labor Actual', 0) or job.get('Sub Labor', 0)
        estimated_labor = job.get('Total Subcontract Est', 0)
        
        # Calculate percent complete (avoid division by zero, cap at 100%)
        if estimated_labor > 0:
            percent_complete = min((labor_actual / estimated_labor) * 100, 100.0)
        else:
            percent_complete = 0.0
        
        labor_data.append({
            'Job Number': job.get('Job Number', ''),
            'Job Description': job.get('Job Name', job.get('Job Description', '')),
            'Contract Amount': job.get('Original Contract Amount', 0),  # Using actual column name
            'Estimated Sub Labor Costs': estimated_labor,  # Using actual column name
            'Monthly Sub Labor Costs': labor_actual,
            'Percent Complete': percent_complete,  # New column
            'Amount Billed': job.get('4020', 0)  # Using 4020 account data for billing
        })
    
    labor_df = pd.DataFrame(labor_data)
    
    # Convert to numeric and filter out rows where Monthly Sub Labor Costs is 0 or blank (include negative values)
    labor_df['Monthly Sub Labor Costs'] = pd.to_numeric(labor_df['Monthly Sub Labor Costs'], errors='coerce').fillna(0)
    labor_df = labor_df[labor_df['Monthly Sub Labor Costs'] != 0]
    
    # 5030 Section - Material Report (4 fields only)
    material_data = []
    for _, job in merged_df.iterrows():
        material_actual = job.get('5030', 0) or job.get('Material Actual', 0) or job.get('Material', 0)
        
        material_data.append({
            'Job Number': job.get('Job Number', ''),
            'Job Description': job.get('Job Name', job.get('Job Description', '')),
            'Estimated Material Costs': job.get('Total Material Estimate', 0),  # Using actual column name
            'Monthly Material Costs': material_actual
        })
    
    material_df = pd.DataFrame(material_data)
    
    # Convert to numeric and filter out rows where Monthly Material Costs is 0 or blank (include negative values)
    material_df['Monthly Material Costs'] = pd.to_numeric(material_df['Monthly Material Costs'], errors='coerce').fillna(0)
    material_df = material_df[material_df['Monthly Material Costs'] != 0]
    
    return labor_df, material_df

def create_excel_update_report(labor_df, material_df):
    """Create a comprehensive Excel report with all updates"""
    
    buffer = io.BytesIO()
    
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        # Labor section updates
        labor_df.to_excel(writer, sheet_name='5040_Labor_Updates', index=False)
        
        # Material section updates
        material_df.to_excel(writer, sheet_name='5030_Material_Updates', index=False)
        
        # Auto-adjust column widths and apply formatting
        from openpyxl.styles import NamedStyle
        
        # Create currency style
        currency_style = NamedStyle(name="currency")
        currency_style.number_format = '$#,##0.00'
        
        # Create percentage style
        percentage_style = NamedStyle(name="percentage")
        percentage_style.number_format = '0.00%'
        
        # Currency columns for each sheet
        currency_columns = {
            '5040_Labor_Updates': ['Contract Amount', 'Monthly Sub Labor Costs', 'Estimated Sub Labor Costs', 'Amount Billed'],
            '5030_Material_Updates': ['Monthly Material Costs', 'Estimated Material Costs']
        }
        
        # Percentage columns for each sheet
        percentage_columns = {
            '5040_Labor_Updates': ['Percent Complete'],
            '5030_Material_Updates': []
        }
        
        for sheet_name in ['5040_Labor_Updates', '5030_Material_Updates']:
            worksheet = writer.sheets[sheet_name]
            
            # Get header row to find column positions
            headers = [cell.value for cell in worksheet[1]]
            
            # Apply currency formatting to appropriate columns
            for col_name in currency_columns[sheet_name]:
                if col_name in headers:
                    col_index = headers.index(col_name) + 1  # Excel is 1-indexed
                    col_letter = worksheet.cell(row=1, column=col_index).column_letter
                    
                    # Apply currency formatting to the entire column (skip header)
                    for row in range(2, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row, column=col_index)
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.number_format = '$#,##0.00'
            
            # Apply percentage formatting to appropriate columns
            for col_name in percentage_columns[sheet_name]:
                if col_name in headers:
                    col_index = headers.index(col_name) + 1  # Excel is 1-indexed
                    col_letter = worksheet.cell(row=1, column=col_index).column_letter
                    
                    # Apply percentage formatting to the entire column (skip header)
                    for row in range(2, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row, column=col_index)
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            # Convert decimal to percentage (divide by 100 since we already multiplied by 100)
                            cell.value = cell.value / 100
                            cell.number_format = '0.00%'
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Add padding, max 50 chars
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Summary sheet
        summary_data = {
            'Section': ['5040 - Labor', '5030 - Material', 'Total'],
            'Jobs Count': [len(labor_df), len(material_df), len(labor_df)],
            'Total Contract Amount': [
                labor_df['Contract Amount'].sum(),
                0,  # Materials don't have contract amount
                labor_df['Contract Amount'].sum()
            ],
            'Total Actual': [
                labor_df['Monthly Sub Labor Costs'].sum(), 
                material_df['Monthly Material Costs'].sum(),
                labor_df['Monthly Sub Labor Costs'].sum() + material_df['Monthly Material Costs'].sum()
            ],
            'Total Budget': [
                labor_df['Estimated Sub Labor Costs'].sum(),
                material_df['Estimated Material Costs'].sum(), 
                labor_df['Estimated Sub Labor Costs'].sum() + material_df['Estimated Material Costs'].sum()
            ],
            'Total Variance': [
                labor_df['Monthly Sub Labor Costs'].sum() - labor_df['Estimated Sub Labor Costs'].sum(),
                material_df['Monthly Material Costs'].sum() - material_df['Estimated Material Costs'].sum(),
                (labor_df['Monthly Sub Labor Costs'].sum() - labor_df['Estimated Sub Labor Costs'].sum()) +
                (material_df['Monthly Material Costs'].sum() - material_df['Estimated Material Costs'].sum())
            ],
            'Total Amount Billed': [
                labor_df['Amount Billed'].sum(),
                0,  # Only labor section has amount billed
                labor_df['Amount Billed'].sum()
            ]
        }
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Apply currency formatting to Summary sheet
        summary_sheet = writer.sheets['Summary']
        summary_headers = [cell.value for cell in summary_sheet[1]]
        
        # Currency columns in summary (all except 'Section' and 'Jobs Count')
        summary_currency_cols = ['Total Contract Amount', 'Total Actual', 'Total Budget', 'Total Variance', 'Total Amount Billed']
        
        for col_name in summary_currency_cols:
            if col_name in summary_headers:
                col_index = summary_headers.index(col_name) + 1
                for row in range(2, summary_sheet.max_row + 1):
                    cell = summary_sheet.cell(row=row, column=col_index)
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        cell.number_format = '$#,##0.00'
        
        # Auto-adjust Summary sheet column widths
        for column in summary_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            summary_sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Instructions sheet
        instructions = [
            "WIP REPORT UPDATE INSTRUCTIONS",
            "",
            "This report contains all the updates for your WIP Report without modifying the original file.",
            "This approach preserves ALL formulas, formatting, and macros in your Excel file.",
            "",
            "HOW TO USE:",
            "",
            "1. LABOR SECTION (5040):",
            "   - Open the '5040_Labor_Updates' tab in this report",
            "   - Copy the 'Monthly Sub Labor Costs' column values", 
            "   - Paste them into the appropriate column in your WIP Report's 5040 section",
            "",
            "2. MATERIAL SECTION (5030):",
            "   - Open the '5030_Material_Updates' tab in this report",
            "   - Copy the 'Monthly Material Costs' column values",
            "   - Paste them into the appropriate column in your WIP Report's 5030 section", 
            "",
            "3. VERIFICATION:",
            "   - Check the 'Summary' tab for totals and variance analysis",
            "   - Variances > $1,000 should be reviewed",
            "",
            "ADVANTAGES OF THIS APPROACH:",
            "✅ NO risk of corrupting your Excel file",
            "✅ ALL formulas and formatting preserved", 
            "✅ All macros and VBA code remain intact",
            "✅ You maintain full control over what gets updated",
            "✅ Easy to verify changes before applying them",
            "",
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ]
        
        instructions_df = pd.DataFrame({'Instructions': instructions})
        instructions_df.to_excel(writer, sheet_name='Instructions', index=False)
        
        # Auto-adjust Instructions sheet column width
        instructions_sheet = writer.sheets['Instructions']
        instructions_sheet.column_dimensions['A'].width = 80  # Wide enough for instructions text
    
    buffer.seek(0)
    return buffer.getvalue()

def display_file_upload_section():
    """Display file upload interface"""
    st.markdown("#### 📁 File Upload")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("**Master WIP Report**")
        master_file = st.file_uploader(
            "Upload Master WIP Report",
            type=['xlsx', 'xlsm'],
            key='master_wip'
        )
        if master_file:
            st.success(f"✅ {master_file.name}")
    
    with col2:
        st.markdown("**WIP Worksheet Export**")
        wip_file = st.file_uploader(
            "Upload WIP Worksheet",
            type=['xlsx'],
            key='wip_worksheet'
        )
        if wip_file:
            st.session_state.files_uploaded['wip'] = wip_file.getvalue()
            st.success(f"✅ {wip_file.name}")
    
    with col3:
        st.markdown("**GL Inquiry Export**")
        gl_file = st.file_uploader(
            "Upload GL Inquiry",
            type=['xlsx'],
            key='gl_inquiry'
        )
        if gl_file:
            st.session_state.files_uploaded['gl'] = gl_file.getvalue()
            st.success(f"✅ {gl_file.name}")

def display_sidebar_options():
    """Display processing options in sidebar"""
    st.sidebar.markdown("### ⚙️ Options")
    
    # Month/Year selector with proper dropdown
    st.sidebar.markdown("**Report Period**")
    
    current_year = datetime.now().year
    years = list(range(current_year - 2, current_year + 2))
    months = [
        "Jan", "Feb", "Mar", "Apr", "May", "Jun",
        "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"
    ]
    
    col1, col2 = st.sidebar.columns(2)
    with col1:
        selected_month = st.selectbox("Month", months, index=3)  # Default to Apr
    with col2:
        selected_year = st.selectbox("Year", years, index=len(years)//2)
    
    # Format as MMM YY
    month_year = f"{selected_month} {str(selected_year)[-2:]}"
    
    st.sidebar.markdown("**Processing Settings**")
    include_closed = st.sidebar.checkbox(
        "Include Closed Jobs", 
        value=False,
        help="Check this to include jobs with 'Closed' status in the report. Useful for quarterly reviews."
    )
    
    return include_closed, month_year

def main():
    st.set_page_config(
        page_title="WIP Report Automation",
        page_icon="📊",
        layout="wide"
    )
    
    # Smaller, cleaner title
    st.markdown("# 📊 WIP Report Automation")
    st.markdown("*Generate update reports without modifying Excel files*")
    st.markdown("---")
    
    initialize_session_state()
    
    # Sidebar options
    include_closed, month_year = display_sidebar_options()
    
    # File Upload Section (main content)
    display_file_upload_section()
    
    st.markdown("---")
    
    # Process Button - better positioned and styled
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        if st.button("🚀 Generate Update Reports", type="primary", use_container_width=True):
            if len(st.session_state.files_uploaded) >= 2:  # Only need WIP and GL
                
                # Process the data
                merged_df = process_data(
                    st.session_state.files_uploaded['wip'],
                    st.session_state.files_uploaded['gl'],
                    include_closed
                )
                
                if merged_df is not None:
                    st.session_state.merged_data = merged_df
                    
                    # Generate reports
                    with st.spinner("Generating update reports..."):
                        labor_df, material_df = generate_update_reports(merged_df)
                        
                        # Create Excel report
                        excel_report = create_excel_update_report(labor_df, material_df)
                    
                    # Store results in session state for display
                    st.session_state.results_ready = True
                    st.session_state.labor_df = labor_df
                    st.session_state.material_df = material_df
                    st.session_state.excel_report = excel_report
                    st.session_state.month_year = month_year
                    
            else:
                st.error("❌ Please upload at least the WIP Worksheet and GL Inquiry files")
    
    # Download button - appears right after processing
    if st.session_state.get('results_ready', False):
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.download_button(
                label="📥 Download Update Reports (Excel)",
                data=st.session_state.excel_report,
                file_name=f"WIP_Update_Reports_{st.session_state.month_year.replace(' ', '')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Download comprehensive update reports that you can use to manually update your WIP Excel file",
                use_container_width=True
            )
    
    # Results Section - Same width as file upload section above
    if st.session_state.get('results_ready', False):
        st.markdown("---")
        
        # Create balanced full-width layout (same as file upload section)
        col1, col2 = st.columns([1, 1])
        
        with col1:
            # Processing Status Section
            st.markdown("### 🚀 Generate Update Reports")
            
            # Show processing status
            st.success(f"✅ Processed {st.session_state.gl_entries} GL entries")
            st.success(f"✅ Merged data for {len(st.session_state.merged_data)} jobs")
            st.success("✅ Update reports generated successfully!")
        
        with col2:
            # Combined Report Summary and Data Preview
            st.markdown("### 📊 Report Summary")
            
            # Calculate variances
            labor_variance = st.session_state.labor_df['Monthly Sub Labor Costs'].sum() - st.session_state.labor_df['Estimated Sub Labor Costs'].sum()
            material_variance = st.session_state.material_df['Monthly Material Costs'].sum() - st.session_state.material_df['Estimated Material Costs'].sum()
            total_variance = labor_variance + material_variance
            
            # Create a clean summary table
            summary_data = {
                'Category': ['Jobs Processed', 'Labor Actual', 'Material Actual', 'Labor Variance', 'Material Variance', 'Total Variance'],
                'Value': [
                    f"{len(st.session_state.merged_data)} jobs",
                    f"${st.session_state.labor_df['Monthly Sub Labor Costs'].sum():,.2f}",
                    f"${st.session_state.material_df['Monthly Material Costs'].sum():,.2f}",
                    f"${labor_variance:,.2f}",
                    f"${material_variance:,.2f}",
                    f"${total_variance:,.2f}"
                ]
            }
            
            summary_df = pd.DataFrame(summary_data)
            st.dataframe(summary_df, use_container_width=True, hide_index=True)
        
        # Data Preview Section - Full width underneath
        st.markdown("---")
        st.markdown("### 📋 Data Preview")
        
        tab1, tab2 = st.tabs(["🔧 5040 - Labor Updates", "📦 5030 - Material Updates"])
        
        with tab1:
            st.markdown("**Labor Section Data (Non-Zero Values Only)**")
            if len(st.session_state.labor_df) > 0:
                st.dataframe(
                    st.session_state.labor_df, 
                    use_container_width=True,
                    height=450
                )
            else:
                st.info("No labor entries with non-zero values found.")
        
        with tab2:
            st.markdown("**Material Section Data (Non-Zero Values Only)**") 
            if len(st.session_state.material_df) > 0:
                st.dataframe(
                    st.session_state.material_df, 
                    use_container_width=True,
                    height=450
                )
            else:
                st.info("No material entries with non-zero values found.")

if __name__ == "__main__":
    main() 