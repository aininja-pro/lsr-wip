import streamlit as st
import pandas as pd
import sys
import os
from datetime import datetime
from pathlib import Path
import io
import numpy as np
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import logging

# Add the src directory to the path so we can import our modules
sys.path.append(str(Path(__file__).parent.parent))

from data_processing.aggregation import aggregate_gl_data
from data_processing.merge_data import process_wip_merge, compute_variances
from data_processing.column_mapping import map_dataframe_columns
from data_processing.excel_integration_v2 import (
    load_wip_workbook, find_or_create_monthly_tab, find_section_markers,
    create_backup, update_wip_report_v2
)

# Configure Streamlit page
st.set_page_config(
    page_title="WIP Report Automation",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Custom CSS for better styling
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: 600;
        color: #2F80ED;
        text-align: center;
        margin-bottom: 2rem;
    }
    .section-header {
        font-size: 1.5rem;
        font-weight: 500;
        color: #333333;
        margin-top: 2rem;
        margin-bottom: 1rem;
    }
    .success-box {
        background-color: #D4EDDA;
        border: 1px solid #C3E6CB;
        border-radius: 0.375rem;
        padding: 1rem;
        margin: 1rem 0;
    }
    .warning-box {
        background-color: #FFF3CD;
        border: 1px solid #FFEAA7;
        border-radius: 0.375rem;
        padding: 1rem;
        margin: 1rem 0;
    }
    .error-box {
        background-color: #F8D7DA;
        border: 1px solid #F5C6CB;
        border-radius: 0.375rem;
        padding: 1rem;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)

def initialize_session_state():
    """Initialize session state variables"""
    if 'files_uploaded' not in st.session_state:
        st.session_state.files_uploaded = {
            'gl_inquiry': None,
            'wip_worksheet': None,
            'master_report': None
        }
    if 'processed_data' not in st.session_state:
        st.session_state.processed_data = None
    if 'processing_complete' not in st.session_state:
        st.session_state.processing_complete = False
    if 'backup_created' not in st.session_state:
        st.session_state.backup_created = None

def display_file_upload_section():
    """Display file upload widgets and validation"""
    st.markdown('<div class="section-header">📁 Upload Required Files</div>', unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.subheader("GL Inquiry Export")
        gl_file = st.file_uploader(
            "Choose GL Inquiry file",
            type=['xlsx', 'xls'],
            key="gl_inquiry_upload",
            help="Export from Sage containing actual costs (accounts 5040, 5030, 4020)"
        )
        if gl_file:
            st.session_state.files_uploaded['gl_inquiry'] = gl_file
            st.success("✓ GL Inquiry uploaded")
    
    with col2:
        st.subheader("WIP Worksheet Export")
        wip_file = st.file_uploader(
            "Choose WIP Worksheet file",
            type=['xlsx', 'xls'],
            key="wip_worksheet_upload",
            help="Export from Sage containing job budgets and status"
        )
        if wip_file:
            st.session_state.files_uploaded['wip_worksheet'] = wip_file
            st.success("✓ WIP Worksheet uploaded")
    
    with col3:
        st.subheader("Master WIP Report")
        master_file = st.file_uploader(
            "Choose Master WIP Report file",
            type=['xlsx', 'xlsm'],
            key="master_report_upload",
            help="The Excel file to be updated (will be backed up first)"
        )
        if master_file:
            st.session_state.files_uploaded['master_report'] = master_file
            st.success("✓ Master Report uploaded")
    
    # Show upload status
    files_ready = all(st.session_state.files_uploaded.values())
    if files_ready:
        st.markdown('<div class="success-box">✅ All required files uploaded and ready for processing!</div>', 
                   unsafe_allow_html=True)
    else:
        missing_files = [name for name, file in st.session_state.files_uploaded.items() if file is None]
        st.markdown(f'<div class="warning-box">⚠️ Still need: {", ".join(missing_files)}</div>', 
                   unsafe_allow_html=True)
    
    return files_ready

def display_processing_options():
    """Display processing configuration options"""
    st.markdown('<div class="section-header">⚙️ Processing Options</div>', unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    
    with col1:
        # Month/Year selector
        current_date = datetime.now()
        selected_date = st.date_input(
            "Select Month/Year to Process",
            value=current_date,
            help="Choose the month and year for the WIP report tab"
        )
        month_year = selected_date.strftime("%b %y")  # Format: "Jun 25"
        
        # Include closed jobs option
        include_closed = st.checkbox(
            "Include Closed Jobs",
            value=False,
            help="Check to include jobs with Status='Closed' (useful for quarterly true-ups)"
        )
    
    with col2:
        # Preview option
        preview_only = st.checkbox(
            "Preview Only (Don't Update Files)",
            value=True,
            help="Check to preview changes without updating the Master WIP Report"
        )
        
        # Backup option
        create_backup = st.checkbox(
            "Create Backup Before Update",
            value=True,
            help="Recommended: Creates timestamped backup before making changes"
        )
    
    return {
        'month_year': month_year,
        'include_closed': include_closed,
        'preview_only': preview_only,
        'create_backup': create_backup
    }

def extract_wip_data(wip_file):
    """Extract the exact fields we need from WIP Worksheet"""
    df = pd.read_excel(wip_file)
    
    # Get the columns by position (0-indexed)
    result = pd.DataFrame()
    result['Job Number'] = df.iloc[:, 0].astype(str).str.strip()  # Column A
    result['Job Name'] = df.iloc[:, 1]  # Column B (Job Description)
    result['Contract Amount'] = pd.to_numeric(df.iloc[:, 3], errors='coerce').fillna(0)  # Column D
    result['Estimated Sub Labor'] = pd.to_numeric(df.iloc[:, 5], errors='coerce').fillna(0)  # Column F
    result['Estimated Material'] = pd.to_numeric(df.iloc[:, 6], errors='coerce').fillna(0)  # Column G
    
    # Filter out rows with empty job numbers
    result = result[result['Job Number'].str.len() > 0]
    
    logger.info(f"Extracted WIP data: {len(result)} records with Job Name, Estimated Sub Labor, and Estimated Material")
    return result

def extract_gl_data(gl_file):
    """Extract Labor Actual, Material Actual, and Amount Billed from GL"""
    df = pd.read_excel(gl_file)
    
    # Filter for relevant accounts
    account_filters = ['5040', '5030', '4020']
    mask = df['Account'].astype(str).str.contains('|'.join(account_filters), na=False)
    df = df[mask]
    
    # Clean job numbers
    df['Job Number'] = df['Job Number'].astype(str).str.strip()
    
    # Convert amounts
    df['Debit'] = pd.to_numeric(df['Debit'], errors='coerce').fillna(0)
    df['Credit'] = pd.to_numeric(df['Credit'], errors='coerce').fillna(0)
    df['Amount'] = df['Debit'] + df['Credit']
    
    # Calculate Amount Billed (K + L)
    amount_billed = 0
    if 'K' in df.columns and 'L' in df.columns:
        df['K'] = pd.to_numeric(df['K'], errors='coerce').fillna(0)
        df['L'] = pd.to_numeric(df['L'], errors='coerce').fillna(0)
        amount_billed = df['K'] + df['L']
    df['Amount Billed'] = amount_billed
    
    # Categorize by account type
    def get_account_type(account):
        if '5040' in str(account):
            return 'Labor'
        elif '5030' in str(account):
            return 'Material'
        return 'Other'
    
    df['Account Type'] = df['Account'].apply(get_account_type)
    
    # Aggregate by Job Number and Account Type
    labor_data = df[df['Account Type'] == 'Labor'].groupby('Job Number').agg({
        'Amount': 'sum',
        'Amount Billed': 'sum'
    }).reset_index()
    labor_data.rename(columns={'Amount': 'Labor Actual'}, inplace=True)
    
    material_data = df[df['Account Type'] == 'Material'].groupby('Job Number').agg({
        'Amount': 'sum'
    }).reset_index()
    material_data.rename(columns={'Amount': 'Material Actual'}, inplace=True)
    
    return labor_data, material_data

def create_final_data(wip_data, labor_data, material_data, include_closed=False):
    """Combine all data for final output"""
    
    # Filter out closed jobs if requested
    if not include_closed and 'Status' in wip_data.columns:
        wip_data = wip_data[wip_data['Status'] != 'Closed']
    
    # Merge with labor data
    final_5040 = wip_data.merge(labor_data, on='Job Number', how='left')
    final_5040['Labor Actual'] = final_5040['Labor Actual'].fillna(0)
    final_5040['Amount Billed'] = final_5040['Amount Billed'].fillna(0)
    
    # Merge with material data  
    final_5030 = wip_data.merge(material_data, on='Job Number', how='left')
    final_5030['Material Actual'] = final_5030['Material Actual'].fillna(0)
    
    return final_5040, final_5030

def update_excel_simple(temp_path, merged_df, progress_callback=None):
    """Write EXACTLY the fields requested to Excel"""
    try:
        # Load workbook
        wb = load_workbook(temp_path, keep_vba=True)
        ws = wb.active
        
        # Debug: Print actual column names
        logger.info(f"Merged DF columns: {list(merged_df.columns)}")
        
        # Find 5040 and 5030 sections
        section_5040_row = None
        section_5030_row = None
        
        for row in range(1, min(50, ws.max_row + 1)):
            for col in range(1, 4):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and '5040' in str(cell_value):
                    section_5040_row = row
                    logger.info(f"Found 5040 section at row {row}")
                if cell_value and '5030' in str(cell_value):
                    section_5030_row = row
                    logger.info(f"Found 5030 section at row {row}")
        
        if not section_5040_row or not section_5030_row:
            raise ValueError("Could not find 5040 or 5030 sections")
        
        # Use simple data access - just write all jobs with the data we have
        logger.info(f"Writing data for {len(merged_df)} jobs")
        
        # Write 5040 section data (Sub Labor jobs only)
        labor_jobs = merged_df[merged_df['Sub Labor Actual'] > 0].copy()
        start_row = section_5040_row + 1
        
        for idx, (_, row_data) in enumerate(labor_jobs.iterrows()):
            excel_row = start_row + idx
            
            # Column A: Job Number (use first column which should be Job Number)
            job_num = str(row_data.iloc[0]).strip() if len(row_data) > 0 else ""
            ws.cell(row=excel_row, column=1).value = job_num
            
            # Column B: Job Description (use second column which should be Job Description)
            job_desc = str(row_data.iloc[1]) if len(row_data) > 1 else ""
            ws.cell(row=excel_row, column=2).value = job_desc
            
            # Column C: Estimated Sub Labor Costs
            estimated_labor = float(row_data.get('Estimated Sub Labor', 0))
            ws.cell(row=excel_row, column=3).value = estimated_labor
            
            # Column D: Monthly Sub Labor Costs (GL aggregation)
            actual_labor = float(row_data.get('Sub Labor Actual', 0))
            ws.cell(row=excel_row, column=4).value = actual_labor
        
        logger.info(f"Wrote {len(labor_jobs)} labor records to 5040 section")
        
        # Write 5030 section data (Material jobs only)
        material_jobs = merged_df[merged_df['Material Actual'] > 0].copy()
        start_row = section_5030_row + 1
        
        for idx, (_, row_data) in enumerate(material_jobs.iterrows()):
            excel_row = start_row + idx
            
            # Column A: Job Number (use first column)
            job_num = str(row_data.iloc[0]).strip() if len(row_data) > 0 else ""
            ws.cell(row=excel_row, column=1).value = job_num
            
            # Column B: Job Description (use second column)
            job_desc = str(row_data.iloc[1]) if len(row_data) > 1 else ""
            ws.cell(row=excel_row, column=2).value = job_desc
            
            # Column C: Estimated Material Costs
            estimated_material = float(row_data.get('Estimated Material', 0))
            ws.cell(row=excel_row, column=3).value = estimated_material
            
            # Column D: Monthly Material Costs (GL aggregation)
            actual_material = float(row_data.get('Material Actual', 0))
            ws.cell(row=excel_row, column=4).value = actual_material
        
        logger.info(f"Wrote {len(material_jobs)} material records to 5030 section")
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Successfully wrote data to Excel")
        return output.getvalue()
        
    except Exception as e:
        logger.error(f"Error updating Excel: {str(e)}")
        raise

def display_data_preview(merged_df, gl_aggregated):
    """Display preview of processed data"""
    st.markdown('<div class="section-header">👁️ Data Preview</div>', unsafe_allow_html=True)
    
    # Summary statistics
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        total_jobs = len(merged_df)
        st.metric("Total Jobs", total_jobs)
    
    with col2:
        jobs_with_activity = len(merged_df[(merged_df['Sub Labor'] > 0) | (merged_df['Material'] > 0)])
        st.metric("Jobs with Activity", jobs_with_activity)
    
    with col3:
        large_variances = len(merged_df[
            (abs(merged_df.get('Sub Labor Variance', 0)) > 1000) |
            (abs(merged_df.get('Material Variance', 0)) > 1000)
        ])
        st.metric("Large Variances (>$1,000)", large_variances)
    
    with col4:
        closed_jobs = len(merged_df[merged_df['Status'] == 'Closed'])
        st.metric("Closed Jobs", closed_jobs)
    
    # Tabbed data display
    tab1, tab2, tab3 = st.tabs(["📊 Merged Data", "🔢 GL Aggregated", "⚠️ Large Variances"])
    
    with tab1:
        st.subheader("WIP Data Merged with GL Actuals")
        st.dataframe(
            merged_df,
            use_container_width=True,
            hide_index=True
        )
    
    with tab2:
        st.subheader("GL Data Aggregated by Job")
        st.dataframe(
            gl_aggregated,
            use_container_width=True,
            hide_index=True
        )
    
    with tab3:
        # Filter for large variances
        variance_mask = (
            (abs(merged_df.get('Sub Labor Variance', 0)) > 1000) |
            (abs(merged_df.get('Material Variance', 0)) > 1000)
        )
        large_variance_df = merged_df[variance_mask]
        
        if len(large_variance_df) > 0:
            st.subheader("Jobs with Variances > $1,000")
            st.dataframe(
                large_variance_df,
                use_container_width=True,
                hide_index=True
            )
        else:
            st.info("No jobs with variances exceeding $1,000")

def display_excel_preview(options):
    """Display preview of Excel sections that would be updated"""
    if not options['preview_only']:
        return
        
    st.markdown('<div class="section-header">📋 Excel Preview</div>', unsafe_allow_html=True)
    
    try:
        # Load the master workbook
        master_file_bytes = st.session_state.files_uploaded['master_report'].getvalue()
        
        with st.spinner("Analyzing Master WIP Report structure..."):
            # Save to temporary file for openpyxl
            temp_path = "temp_master.xlsx"
            with open(temp_path, "wb") as f:
                f.write(master_file_bytes)
            
            # Load workbook and find sections
            wb = load_wip_workbook(temp_path)
            ws = find_or_create_monthly_tab(wb, options['month_year'])
            
            # Find sections
            section_markers = find_section_markers(ws, ["5040", "5030"])
            section_5040_row = section_markers.get("5040", (None, None))[0] if section_markers.get("5040") else None
            section_5030_row = section_markers.get("5030", (None, None))[0] if section_markers.get("5030") else None
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.subheader("🔵 5040 Section (Sub Labor)")
                if section_5040_row:
                    st.success(f"✓ Found at row {section_5040_row}")
                    st.info("Section detected and ready for updates")
                else:
                    st.error("❌ Section not found")
            
            with col2:
                st.subheader("🟢 5030 Section (Material)")
                if section_5030_row:
                    st.success(f"✓ Found at row {section_5030_row}")
                    st.info("Section detected and ready for updates")
                else:
                    st.error("❌ Section not found")
            
            # Clean up temp file
            os.remove(temp_path)
            
    except Exception as e:
        st.error(f"Error analyzing Excel structure: {str(e)}")

def update_excel_file(merged_df, options):
    """Update the Excel file with processed data"""
    if options['preview_only']:
        st.info("Preview mode - no files will be updated")
        return None
        
    try:
        with st.spinner("Updating Master WIP Report..."):
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            # Save master file temporarily
            master_file_bytes = st.session_state.files_uploaded['master_report'].getvalue()
            temp_path = "/app/temp_master_update.xlsx"
            with open(temp_path, "wb") as f:
                f.write(master_file_bytes)
            
            status_text.text("📂 Processing data...")
            progress_bar.progress(40)
            
            # Create backup if requested
            if options['create_backup']:
                backup_dir = "/app/WIP_Backups"
                os.makedirs(backup_dir, exist_ok=True)
                backup_filename = f"WIP_Report_BACKUP_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                backup_path = os.path.join(backup_dir, backup_filename)
                with open(backup_path, "wb") as f:
                    f.write(master_file_bytes)
                st.success(f"Backup created: {backup_filename}")
                st.session_state.backup_created = backup_filename
            
            # Extract WIP data for the fields we need
            wip_file_bytes = st.session_state.files_uploaded['wip_worksheet'].getvalue()
            wip_data = extract_wip_data(wip_file_bytes)
            
            # Extract GL data for the fields we need
            gl_file_bytes = st.session_state.files_uploaded['gl_inquiry'].getvalue()
            labor_data, material_data = extract_gl_data(gl_file_bytes)
            
            # Create final datasets with all the fields
            data_5040, data_5030 = create_final_data(wip_data, labor_data, material_data, 
                                                   include_closed=options['include_closed'])
            
            status_text.text("✍️ Updating Excel file...")
            progress_bar.progress(80)
            
            # Use the correct update function that writes all fields
            updated_bytes = update_excel_simple(temp_path, merged_df, progress_callback=progress_bar.progress)
            
            progress_bar.progress(100)
            status_text.text("✅ Update complete!")
            
            # Clean up
            os.remove(temp_path)
            
            return updated_bytes
            
    except Exception as e:
        st.error(f"Error updating Excel file: {str(e)}")
        if os.path.exists(temp_path):
            os.remove(temp_path)
        return None

def display_download_section(updated_file_bytes, merged_df):
    """Display download buttons for updated files"""
    st.markdown('<div class="section-header">📥 Download Results</div>', unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    
    with col1:
        if updated_file_bytes:
            st.download_button(
                label="📊 Download Updated WIP Report",
                data=updated_file_bytes,
                file_name=f"WIP_Report_Updated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Download the updated Master WIP Report"
            )
        else:
            st.info("Updated report will appear here after processing")
    
    with col2:
        # Create validation report
        if merged_df is not None:
            validation_df = merged_df[
                (abs(merged_df.get('Sub Labor Variance', 0)) > 1000) |
                (abs(merged_df.get('Material Variance', 0)) > 1000)
            ].copy()
            
            if not validation_df.empty:
                # Convert to Excel bytes with proper buffer handling
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    validation_df.to_excel(writer, sheet_name='Large Variances', index=False)
                
                # Critical: seek to beginning before getting value
                output.seek(0)
                validation_bytes = output.getvalue()
                output.close()
                
                st.download_button(
                    label="⚠️ Download Validation Report",
                    data=validation_bytes,
                    file_name=f"WIP_Validation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="Download report of jobs with large variances"
                )
            else:
                st.success("No large variances found - no validation report needed")

def load_and_process_gl_data(gl_file):
    """Load and process GL data for display"""
    df = pd.read_excel(gl_file)
    
    # Find the actual column names (flexible mapping)
    actual_columns = df.columns.tolist()
    logger.info(f"Available columns in GL file: {actual_columns}")
    
    # Find Job Number column
    job_col = None
    for col in actual_columns:
        if any(variant.lower() in col.lower() for variant in ['job number', 'job no', 'job #', 'job']):
            job_col = col
            break
    
    if not job_col:
        raise ValueError(f"Could not find Job Number column. Available columns: {actual_columns}")
    
    # Find Account column
    account_col = None
    for col in actual_columns:
        if any(variant.lower() in col.lower() for variant in ['account', 'acct', 'gl account']):
            account_col = col
            break
    
    if not account_col:
        raise ValueError(f"Could not find Account column. Available columns: {actual_columns}")
    
    # Filter for relevant accounts
    account_filters = ['5040', '5030', '4020']
    mask = df[account_col].astype(str).str.contains('|'.join(account_filters), na=False)
    df = df[mask]
    logger.info(f"Filtered GL data from {len(df)} to {len(df)} records based on account filters: {account_filters}")
    
    # Clean job numbers
    df['Job Number'] = df[job_col].astype(str).str.strip()
    
    # Find Debit and Credit columns
    debit_col = None
    credit_col = None
    for col in actual_columns:
        if any(variant.lower() in col.lower() for variant in ['debit', 'dr', 'debit amount']):
            debit_col = col
        if any(variant.lower() in col.lower() for variant in ['credit', 'cr', 'credit amount']):
            credit_col = col
    
    # Convert amounts
    df['Debit'] = pd.to_numeric(df[debit_col], errors='coerce').fillna(0) if debit_col else 0
    df['Credit'] = pd.to_numeric(df[credit_col], errors='coerce').fillna(0) if credit_col else 0
    
    # Calculate Amount Billed (K + L) if columns exist
    if 'K' in df.columns and 'L' in df.columns:
        df['K'] = pd.to_numeric(df['K'], errors='coerce').fillna(0)
        df['L'] = pd.to_numeric(df['L'], errors='coerce').fillna(0)
        logger.info("Found Amount Billed columns (K, L)")
    else:
        logger.warning("Amount Billed columns (K, L) not found, setting Amount Billed to 0")
    
    df['Amount'] = df['Debit'] + df['Credit']
    logger.info("Computed Amount field as Debit + Credit")
    
    # Categorize by account type
    def get_account_type(account):
        if '5040' in str(account):
            return 'Sub Labor'
        elif '5030' in str(account):
            return 'Material'
        return 'Other'
    
    df['Account Type'] = df[account_col].apply(get_account_type)
    
    # Aggregate by Job Number and Account Type
    aggregated = df.groupby(['Job Number', 'Account Type']).agg({
        'Amount': 'sum'
    }).reset_index()
    
    logger.info(f"Aggregated GL data to {len(aggregated)} job records")
    return aggregated

def load_wip_worksheet(wip_file):
    """Load WIP worksheet for display"""
    df = pd.read_excel(wip_file)
    
    # Extract the fields we need using column positions
    result = pd.DataFrame()
    result['Job Number'] = df.iloc[:, 0].astype(str).str.strip()  # Column A
    result['Job Description'] = df.iloc[:, 1]  # Column B
    result['Status'] = df.iloc[:, 2] if df.shape[1] > 2 else 'Active'  # Column C
    result['Contract Amount'] = pd.to_numeric(df.iloc[:, 3], errors='coerce').fillna(0)  # Column D
    result['Estimated Sub Labor'] = pd.to_numeric(df.iloc[:, 5], errors='coerce').fillna(0)  # Column F
    result['Estimated Material'] = pd.to_numeric(df.iloc[:, 6], errors='coerce').fillna(0)  # Column G
    
    return result

def merge_data(wip_df, gl_df, include_closed=False):
    """Merge WIP and GL data"""
    # Filter out closed jobs if requested
    if not include_closed and 'Status' in wip_df.columns:
        wip_df = wip_df[wip_df['Status'] != 'Closed']
    
    # Pivot GL data to get Sub Labor and Material columns
    gl_pivot = gl_df.pivot_table(
        index='Job Number', 
        columns='Account Type', 
        values='Amount', 
        fill_value=0
    ).reset_index()
    
    # Merge with WIP data
    merged = wip_df.merge(gl_pivot, on='Job Number', how='left')
    
    # Fill missing values
    for col in ['Sub Labor', 'Material']:
        if col in merged.columns:
            merged[col] = merged[col].fillna(0)
        else:
            merged[col] = 0
    
    # Rename for clarity
    merged.rename(columns={
        'Sub Labor': 'Sub Labor Actual',
        'Material': 'Material Actual'
    }, inplace=True)
    
    # Add Amount Billed as 0 for now (will be filled by extract_gl_data)
    merged['Amount Billed'] = 0
    
    logger.info(f"Merged WIP data ({len(wip_df)} records) with GL data ({len(gl_df)} records). Result: {len(merged)} records")
    return merged

def main():
    """Main Streamlit application"""
    initialize_session_state()
    
    # Header
    st.markdown('<div class="main-header">WIP Report Automation Tool</div>', unsafe_allow_html=True)
    st.markdown("---")
    
    # Sidebar with instructions
    with st.sidebar:
        st.header("📖 Instructions")
        st.markdown("""
        **Step 1:** Upload all three required Excel files
        
        **Step 2:** Configure processing options
        
        **Step 3:** Click Process to run automation
        
        **Step 4:** Review preview and download results
        
        ---
        
        **Files Required:**
        - **GL Inquiry**: Actual costs from Sage
        - **WIP Worksheet**: Job budgets from Sage  
        - **Master WIP Report**: File to be updated
        
        ---
        
        **Safety Features:**
        - ✅ Automatic backups
        - ✅ Formula preservation
        - ✅ Preview before update
        - ✅ Validation reports
        """)
    
    # Main content
    files_ready = display_file_upload_section()
    
    if files_ready:
        options = display_processing_options()
        
        # Process button
        if st.button("🚀 Process Data", type="primary", use_container_width=True):
            try:
                with st.spinner("Processing data..."):
                    # Load and process data
                    st.info("Loading GL Inquiry data...")
                    gl_df = load_and_process_gl_data(st.session_state.files_uploaded['gl_inquiry'])
                    st.success(f"Processed {len(gl_df)} GL records")
                    
                    st.info("Loading WIP Worksheet data...")
                    wip_df = load_wip_worksheet(st.session_state.files_uploaded['wip_worksheet'])
                    st.success(f"Loaded {len(wip_df)} WIP records")
                    
                    st.info("Merging data...")
                    merged_df = merge_data(wip_df, gl_df, include_closed=options['include_closed'])
                    st.success(f"Merged to {len(merged_df)} final records")
                    
                    # Display preview with ALL the new fields
                    st.subheader("📊 Data Preview - All Required Fields")
                    preview_cols = ['Job Number', 'Job Description', 'Status', 
                                  'Contract Amount', 'Estimated Sub Labor', 'Sub Labor Actual',
                                  'Estimated Material', 'Material Actual', 'Amount Billed']
                    
                    display_df = merged_df[preview_cols].copy()
                    for col in ['Contract Amount', 'Estimated Sub Labor', 'Sub Labor Actual', 
                               'Estimated Material', 'Material Actual', 'Amount Billed']:
                        if col in display_df.columns:
                            display_df[col] = display_df[col].apply(lambda x: f"${x:,.2f}" if pd.notna(x) else "$0.00")
                    
                    st.dataframe(display_df, use_container_width=True)
                    
                    # Update Excel file
                    if not options['preview_only']:
                        st.info("Updating WIP Report...")
                        updated_bytes = update_excel_file(merged_df, options)
                        
                        if updated_bytes:
                            st.success("✅ WIP Report updated successfully!")
                            
                            # Download button
                            st.download_button(
                                label="📊 Download Updated WIP Report",
                                data=updated_bytes,
                                file_name=f"WIP_Report_{options['month_year'].replace(' ', '')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                            
                            # Summary
                            st.subheader("📈 Processing Summary")
                            col1, col2, col3, col4, col5 = st.columns(5)
                            with col1:
                                st.metric("Jobs Processed", len(merged_df))
                            with col2:
                                st.metric("Contract Amount", f"${merged_df['Contract Amount'].sum():,.0f}")
                            with col3:
                                st.metric("Sub Labor Actual", f"${merged_df['Sub Labor Actual'].sum():,.0f}")
                            with col4:
                                st.metric("Material Actual", f"${merged_df['Material Actual'].sum():,.0f}")
                            with col5:
                                st.metric("Amount Billed", f"${merged_df['Amount Billed'].sum():,.0f}")
                    else:
                        st.info("Preview mode - Excel file not updated")
                    
            except Exception as e:
                st.error(f"Error processing data: {str(e)}")
                logger.error(f"Processing error: {str(e)}")

if __name__ == "__main__":
    main() 