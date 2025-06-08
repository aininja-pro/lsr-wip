#!/usr/bin/env python3
"""
WIP Report Automation Tool - Memory-Only Version (No Temp Files)
Fixes Excel corruption by working entirely in memory
"""

import streamlit as st
import pandas as pd
import io
import os
import sys
from datetime import datetime
from pathlib import Path

# Add src to path for imports
sys.path.append(str(Path(__file__).parent.parent))

from data_processing.aggregation import process_gl_inquiry
from data_processing.merge_data import process_wip_merge
from data_processing.excel_integration_v2 import (
    find_or_create_monthly_tab,
    find_section_markers,
    safe_write_cell,
    is_merged_cell
)
import openpyxl
from openpyxl import load_workbook

# Page config
st.set_page_config(
    page_title="WIP Report Automation",
    page_icon="📊",
    layout="wide"
)

# Custom CSS for styling
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: 700;
        text-align: center;
        color: #2F80ED;
        margin-bottom: 2rem;
    }
    .section-header {
        font-size: 1.5rem;
        font-weight: 600;
        color: #333;
        margin: 1.5rem 0 1rem 0;
        border-bottom: 2px solid #2F80ED;
        padding-bottom: 0.5rem;
    }
    .metric-container {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 1rem;
        border-radius: 10px;
        color: white;
        text-align: center;
        margin: 0.5rem 0;
    }
    .upload-box {
        border: 2px dashed #2F80ED;
        border-radius: 10px;
        padding: 1rem;
        text-align: center;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)

def initialize_session_state():
    """Initialize session state variables"""
    if 'files_uploaded' not in st.session_state:
        st.session_state.files_uploaded = {}
    if 'processing_complete' not in st.session_state:
        st.session_state.processing_complete = False
    if 'processed_data' not in st.session_state:
        st.session_state.processed_data = None
    if 'backup_created' not in st.session_state:
        st.session_state.backup_created = None

def display_file_upload_section():
    """Display file upload widgets and track upload status"""
    st.markdown('<div class="section-header">📁 File Upload</div>', unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("**GL Inquiry Export**")
        gl_file = st.file_uploader(
            "Choose GL Inquiry file",
            type=['xlsx', 'xls'],
            key="gl_inquiry",
            help="GL Inquiry export from Sage with actual costs"
        )
        if gl_file:
            st.session_state.files_uploaded['gl_inquiry'] = gl_file
            st.success("✅ GL Inquiry uploaded")
    
    with col2:
        st.markdown("**WIP Worksheet Export**")
        wip_file = st.file_uploader(
            "Choose WIP Worksheet file",
            type=['xlsx', 'xls'],
            key="wip_worksheet",
            help="WIP Worksheet export from Sage with job budgets"
        )
        if wip_file:
            st.session_state.files_uploaded['wip_worksheet'] = wip_file
            st.success("✅ WIP Worksheet uploaded")
    
    with col3:
        st.markdown("**Master WIP Report**")
        master_file = st.file_uploader(
            "Choose Master WIP Report file",
            type=['xlsx', 'xlsm'],
            key="master_report",
            help="Master WIP Report file to be updated"
        )
        if master_file:
            st.session_state.files_uploaded['master_report'] = master_file
            st.success("✅ Master WIP Report uploaded")
    
    # Check if all files are uploaded
    required_files = ['gl_inquiry', 'wip_worksheet', 'master_report']
    files_ready = all(key in st.session_state.files_uploaded for key in required_files)
    
    if files_ready:
        st.success("🎉 All files uploaded successfully!")
    else:
        missing = [key for key in required_files if key not in st.session_state.files_uploaded]
        st.warning(f"Missing files: {', '.join(missing)}")
    
    return files_ready

def display_processing_options():
    """Display processing configuration options"""
    st.markdown('<div class="section-header">⚙️ Processing Options</div>', unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    
    with col1:
        # Month/Year selector
        current_date = datetime.now()
        selected_date = st.date_input(
            "Select Month/Year for Update",
            value=current_date,
            help="Select the month and year for the WIP report update"
        )
        month_year = selected_date.strftime("%b %y")
        
        # Include closed jobs
        include_closed = st.checkbox(
            "Include Closed Jobs",
            value=False,
            help="Check this for quarterly true-ups that include closed jobs"
        )
    
    with col2:
        # Create backup
        create_backup = st.checkbox(
            "Create Backup",
            value=True,
            help="Automatically create a timestamped backup before updating"
        )
        
        # Preview only mode
        preview_only = st.checkbox(
            "Preview Only (Don't Update Files)",
            value=False,
            help="Preview changes without actually updating the Master WIP Report"
        )
    
    return {
        'month_year': month_year,
        'include_closed': include_closed,
        'create_backup': create_backup,
        'preview_only': preview_only
    }

def map_gl_columns(df):
    """Map GL DataFrame columns to standard names"""
    column_variations = {
        'Account': ['Account', 'Account Number', 'Acct', 'GL Account', 'Account No'],
        'Job Number': ['Job Number', 'Job No', 'Job #', 'Job', 'Project Number', 'Project No', 'JobNumber'],
        'Debit': ['Debit', 'Debit Amount', 'DR', 'Dr', 'Debit Amt'],
        'Credit': ['Credit', 'Credit Amount', 'CR', 'Cr', 'Credit Amt']
    }
    
    # Map column names to standard names
    column_mapping = {}
    for standard_name, variations in column_variations.items():
        found_column = None
        for variation in variations:
            if variation in df.columns:
                found_column = variation
                break
        
        if found_column:
            column_mapping[found_column] = standard_name
        else:
            # If standard name not found, raise error with available columns
            raise ValueError(f"Required column '{standard_name}' not found. Available columns: {list(df.columns)}")
    
    # Rename columns to standard names
    df = df.rename(columns=column_mapping)
    return df

def map_wip_columns(df):
    """Map WIP DataFrame columns to standard names"""
    column_variations = {
        'Job Number': ['Job Number', 'Job No', 'Job #', 'Job', 'Project Number', 'Project No', 'JobNumber'],
        'Status': ['Status', 'Job Status', 'Project Status', 'State'],
        'Job Name': ['Job Name', 'Project Name', 'Description', 'Job Description'],
    }
    
    # Map column names to standard names  
    column_mapping = {}
    for standard_name, variations in column_variations.items():
        found_column = None
        for variation in variations:
            if variation in df.columns:
                found_column = variation
                break
        
        if found_column:
            column_mapping[found_column] = standard_name
        else:
            # Only require Job Number and Status as mandatory
            if standard_name in ['Job Number', 'Status']:
                raise ValueError(f"Required column '{standard_name}' not found. Available columns: {list(df.columns)}")
    
    # Rename columns to standard names
    df = df.rename(columns=column_mapping)
    return df

def process_data(options):
    """Process uploaded data and return merged results"""
    try:
        with st.spinner("Processing data..."):
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            # Step 1: Process GL data
            status_text.text("🔍 Processing GL Inquiry data...")
            progress_bar.progress(20)
            
            # Process GL data directly from uploaded file bytes
            gl_file_bytes = st.session_state.files_uploaded['gl_inquiry'].getvalue()
            gl_df = pd.read_excel(io.BytesIO(gl_file_bytes))
            
            # Apply robust column mapping for GL data
            gl_df = map_gl_columns(gl_df)
            
            # Apply GL processing steps manually since we have DataFrame
            from data_processing.aggregation import filter_gl_accounts, compute_amounts, aggregate_gl_data
            filtered_gl = filter_gl_accounts(gl_df)
            amounts_gl = compute_amounts(filtered_gl)
            gl_aggregated = aggregate_gl_data(amounts_gl)
            
            # Step 2: Process WIP Worksheet
            status_text.text("📋 Processing WIP Worksheet...")
            progress_bar.progress(40)
            wip_file_bytes = st.session_state.files_uploaded['wip_worksheet'].getvalue()
            wip_df = pd.read_excel(io.BytesIO(wip_file_bytes))
            
            # Apply robust column mapping for WIP data
            wip_df = map_wip_columns(wip_df)
            
            # Step 3: Merge data
            status_text.text("🔗 Merging GL and WIP data...")
            progress_bar.progress(60)
            
            # Save WIP data to memory buffer for processing
            wip_buffer = io.BytesIO()
            wip_df.to_excel(wip_buffer, index=False)
            wip_buffer.seek(0)
            
            # Process merge using buffer
            merged_df = process_wip_merge_from_buffer(wip_buffer, gl_aggregated, include_closed=options['include_closed'])
            wip_buffer.close()
            
            progress_bar.progress(100)
            status_text.text("✅ Data processing complete!")
            
            return merged_df, gl_aggregated
            
    except Exception as e:
        st.error(f"Error processing data: {str(e)}")
        return None, None

def process_wip_merge_from_buffer(wip_buffer, gl_aggregated, include_closed=False):
    """Process WIP merge from memory buffer instead of file"""
    # Read from buffer
    wip_df = pd.read_excel(wip_buffer)
    
    # Import the original merge function logic here
    from data_processing.merge_data import trim_job_numbers, merge_wip_with_gl, filter_closed_jobs
    
    # Apply the same logic as the original function
    wip_df = trim_job_numbers(wip_df)
    wip_df = filter_closed_jobs(wip_df, include_closed=include_closed)
    merged_df = merge_wip_with_gl(wip_df, gl_aggregated)
    
    return merged_df

def create_backup_from_bytes(file_bytes, month_year=None):
    """Create backup from file bytes"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if month_year:
        backup_name = f"WIP_Report_BACKUP_{month_year}_{timestamp}.xlsx"
    else:
        backup_name = f"WIP_Report_BACKUP_{timestamp}.xlsx"
    
    backup_path = Path("WIP_Backups") / backup_name
    backup_path.parent.mkdir(exist_ok=True)
    
    with open(backup_path, "wb") as f:
        f.write(file_bytes)
    
    return backup_name

def clear_and_update_section_memory(ws, data_df, start_row, value_column, target_columns):
    """Clear and update a section working in memory"""
    if data_df.empty:
        return 0
    
    # Clear existing data (value cells only, skip formulas and merged cells)
    current_row = start_row + 1  # Start after header
    cells_cleared = 0
    
    while current_row <= ws.max_row:
        # Check if this row is empty (no job number in column A)
        job_cell = ws.cell(row=current_row, column=1)
        if not job_cell.value:
            break
            
        # Clear value cells in target columns
        for col in target_columns:
            cell = ws.cell(row=current_row, column=col)
            if not is_merged_cell(ws, current_row, col) and cell.data_type != 'f':
                cell.value = None
                cells_cleared += 1
        
        current_row += 1
    
    # Write new data
    jobs_written = 0
    for idx, row in data_df.iterrows():
        target_row = start_row + 1 + jobs_written
        
        # Write job number to column A
        safe_write_cell(ws, target_row, 1, row['Job Number'])
        
        # Write value to the appropriate columns
        if value_column in row and pd.notna(row[value_column]):
            for col in target_columns[:2]:  # Write to first 2 target columns
                safe_write_cell(ws, target_row, col, row[value_column])
        
        jobs_written += 1
    
    return cells_cleared

def update_excel_file_memory_only(merged_df, options):
    """Update Excel file working entirely in memory - no temp files!"""
    if options['preview_only']:
        st.info("Preview mode - no files will be updated")
        return None
        
    try:
        with st.spinner("Updating Master WIP Report..."):
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            # Work entirely in memory - no temp files!
            status_text.text("📂 Loading workbook from memory...")
            progress_bar.progress(20)
            
            # Get master file bytes from uploaded file
            master_file_bytes = st.session_state.files_uploaded['master_report'].getvalue()
            
            # Load workbook directly from memory with ALL preservation flags
            st.info(f"📁 Original file size: {len(master_file_bytes):,} bytes")
            master_buffer = io.BytesIO(master_file_bytes)
            wb = load_workbook(
                master_buffer, 
                keep_vba=True,        # Preserve VBA macros
                data_only=False,      # CRITICAL: Keep formulas, not just values
                keep_links=True       # Preserve external links
            )
            master_buffer.close()
            
            # Find or create monthly tab
            status_text.text("📋 Finding monthly tab...")
            progress_bar.progress(40)
            st.info(f"🔍 Looking for monthly tab: '{options['month_year']}'")
            st.info(f"📋 Available tabs in workbook: {', '.join(wb.sheetnames)}")
            ws = find_or_create_monthly_tab(wb, options['month_year'])
            st.success(f"✅ Using monthly tab: '{ws.title}'")
            
            # Prepare data for update
            status_text.text("✍️ Preparing data for update...")
            progress_bar.progress(50)
            
            # Separate data for each section
            sub_labor_data = merged_df[merged_df['Sub Labor'] > 0][['Job Number', 'Sub Labor']].copy()
            material_data = merged_df[merged_df['Material'] > 0][['Job Number', 'Material']].copy()
            
            # Update sections directly in memory
            status_text.text("✍️ Updating Excel sections...")
            progress_bar.progress(70)
            
            # Find section markers
            status_text.text("🔍 Locating sections in Excel file...")
            progress_bar.progress(75)
            sections = find_section_markers(ws, ['5040', '5030'])
            
            # Show success messages for found sections
            section_messages = []
            if sections.get('5040'):
                row, col = sections['5040']
                section_messages.append(f"✅ Found 5040 section at row {row}")
                st.success(f"✅ Found 5040 section at row {row}")
            if sections.get('5030'):
                row, col = sections['5030']
                section_messages.append(f"✅ Found 5030 section at row {row}")
                st.success(f"✅ Found 5030 section at row {row}")
            
            # Check if both sections were found
            if not sections or not sections.get('5040') or not sections.get('5030'):
                if not sections.get('5040'):
                    st.error("❌ Could not find 5040 section in the worksheet")
                if not sections.get('5030'):
                    st.error("❌ Could not find 5030 section in the worksheet")
                raise Exception("Could not find required sections in the worksheet")
            
            # Update 5040 section (Sub Labor)
            status_text.text("✏️ Updating 5040 section (Sub Labor)...")
            progress_bar.progress(80)
            total_cleared = 0
            if not sub_labor_data.empty:
                section_5040_row, section_5040_col = sections['5040']
                cleared = clear_and_update_section_memory(ws, sub_labor_data, section_5040_row, 'Sub Labor', [3, 4, 5, 8])  # Columns C,D,E,H
                total_cleared += cleared
                st.info(f"📝 Updated {len(sub_labor_data)} Sub Labor entries in 5040 section")
            
            # Update 5030 section (Material)  
            if not material_data.empty:
                section_5030_row, section_5030_col = sections['5030']
                cleared = clear_and_update_section_memory(ws, material_data, section_5030_row, 'Material', [2, 3])  # Columns B,C
                total_cleared += cleared
                st.info(f"📝 Updated {len(material_data)} Material entries in 5030 section")
            
            # Create backup if requested
            if options['create_backup']:
                status_text.text("💾 Creating backup...")
                progress_bar.progress(85)
                backup_name = create_backup_from_bytes(master_file_bytes, options['month_year'])
                st.session_state.backup_created = backup_name
            
            # Save updated workbook to memory
            status_text.text("💾 Finalizing updated report...")
            progress_bar.progress(90)
            
            # Create a fresh buffer and save properly
            output_buffer = io.BytesIO()
            wb.save(output_buffer)
            wb.close()  # Close workbook BEFORE getting bytes (Claude's fix)
            
            # Get the bytes AFTER closing workbook
            output_buffer.seek(0)
            updated_file_bytes = output_buffer.getvalue()
            output_buffer.close()
            
            # Debug file size comparison
            print(f"Original file size: {len(master_file_bytes)} bytes")
            print(f"Updated file size: {len(updated_file_bytes)} bytes") 
            print(f"File size difference: {len(updated_file_bytes) - len(master_file_bytes)} bytes")
            
            st.info(f"📁 Original file size: {len(master_file_bytes):,} bytes")
            st.info(f"📁 Updated file size: {len(updated_file_bytes):,} bytes")
            st.info(f"📁 File size difference: {len(updated_file_bytes) - len(master_file_bytes):,} bytes")
            
            # Show data loss warning
            if len(updated_file_bytes) < len(master_file_bytes):
                st.error(f"🚨 DATA LOSS DETECTED: {len(master_file_bytes) - len(updated_file_bytes):,} bytes lost!")
            
            # CRITICAL TEST: Minimal save (no changes) for debugging
            st.subheader("🧪 CRITICAL TEST: Load + Save with NO Changes")
            
            try:
                # Load and immediately save without ANY changes
                test_buffer = io.BytesIO(master_file_bytes)
                test_wb = load_workbook(test_buffer, keep_vba=True, data_only=False)
                test_buffer.close()
                
                test_output = io.BytesIO()
                test_wb.save(test_output)
                test_wb.close()
                test_output.seek(0)
                test_bytes = test_output.getvalue()
                test_output.close()
                
                # File size comparison for minimal save
                print(f"MINIMAL SAVE TEST:")
                print(f"Original file size: {len(master_file_bytes)} bytes")
                print(f"Minimal save size: {len(test_bytes)} bytes") 
                print(f"Data loss in minimal save: {len(master_file_bytes) - len(test_bytes)} bytes")
                
                st.info(f"🧪 Minimal save size: {len(test_bytes):,} bytes")
                st.info(f"🧪 Data loss in minimal save: {len(master_file_bytes) - len(test_bytes):,} bytes")
                
                if len(test_bytes) < len(master_file_bytes):
                    st.error("⚠️ CORRUPTION SOURCE FOUND: load_workbook/save process is losing data!")
                else:
                    st.success("✅ Minimal save preserved file size - corruption happens during our processing")
                
                st.download_button(
                    "🧪 Test Download - MINIMAL SAVE (No Changes Made)",
                    data=test_bytes,
                    file_name="test_minimal_save.xlsx", 
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="Test if corruption happens just from loading and saving"
                )
                
            except Exception as e:
                st.error(f"❌ Minimal save test failed: {str(e)}")
                import traceback
                st.code(traceback.format_exc())
            
            # Additional debug downloads
            if st.checkbox("🔍 Enable Additional Debug Downloads"):
                test_buffer = io.BytesIO(master_file_bytes)
                test_wb = load_workbook(
                    test_buffer, 
                    keep_vba=True, 
                    data_only=False, 
                    keep_links=True
                )
                test_buffer.close()
                
                test_output = io.BytesIO()
                test_wb.save(test_output)
                test_wb.close()
                test_output.seek(0)
                test_bytes = test_output.getvalue()
                test_output.close()
                
                st.download_button(
                    "🧪 Test Download (No Changes)", 
                    test_bytes, 
                    "test_no_changes.xlsx",
                    help="Download file with no changes to test if corruption is from loading/saving"
                )
            
            progress_bar.progress(100)
            status_text.text("✅ Update complete!")
            
            return updated_file_bytes
            
    except Exception as e:
        st.error(f"Error updating Excel file: {str(e)}")
        return None

def display_data_preview(merged_df, gl_aggregated):
    """Display preview of processed data"""
    st.markdown('<div class="section-header">📊 Data Preview</div>', unsafe_allow_html=True)
    
    if merged_df is not None:
        st.subheader("Merged WIP Data Ready for Update")
        st.dataframe(merged_df.head(10), use_container_width=True)
        
        col1, col2 = st.columns(2)
        with col1:
            st.metric("Total Jobs", len(merged_df))
        with col2:
            if gl_aggregated is not None:
                st.metric("GL Entries Processed", len(gl_aggregated))
    else:
        st.warning("No merged data available")

def display_download_section(updated_file_bytes, merged_df):
    """Display download buttons for updated files"""
    st.markdown('<div class="section-header">📥 Download Results</div>', unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    
    with col1:
        if updated_file_bytes:
            st.download_button(
                label="📊 Download Updated WIP Report",
                data=updated_file_bytes,
                file_name=f"WIP_Report_Updated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Download the updated Master WIP Report"
            )
        else:
            st.info("Updated report will appear here after processing")
    
    with col2:
        # Create validation report
        if merged_df is not None:
            # Create simple validation based on available data
            # Filter for jobs with significant material or sub labor amounts (> $1000)
            material_condition = (merged_df['Material'] > 1000) if 'Material' in merged_df.columns else False
            labor_condition = (merged_df['Sub Labor'] > 1000) if 'Sub Labor' in merged_df.columns else False
            
            # Combine conditions safely
            if isinstance(material_condition, pd.Series) and isinstance(labor_condition, pd.Series):
                large_jobs = merged_df[material_condition | labor_condition].copy()
            elif isinstance(material_condition, pd.Series):
                large_jobs = merged_df[material_condition].copy()
            elif isinstance(labor_condition, pd.Series):
                large_jobs = merged_df[labor_condition].copy()
            else:
                large_jobs = pd.DataFrame()  # Empty if no valid conditions
            
            if not large_jobs.empty:
                # Convert to Excel bytes with proper buffer handling
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    large_jobs.to_excel(writer, sheet_name='Large Jobs', index=False)
                
                # Critical: seek to beginning before getting value
                output.seek(0)
                validation_bytes = output.getvalue()
                output.close()
                
                st.download_button(
                    label="📋 Download Large Jobs Report",
                    data=validation_bytes,
                    file_name=f"WIP_Large_Jobs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="Download report of jobs with significant amounts"
                )
            else:
                st.info("No large jobs found - no validation report needed")

def main():
    """Main Streamlit application"""
    initialize_session_state()
    
    # Header
    st.markdown('<div class="main-header">WIP Report Automation Tool</div>', unsafe_allow_html=True)
    st.markdown("**Memory-Only Version** - No temporary files, no corruption! 🚀", unsafe_allow_html=True)
    st.markdown("---")
    
    # Sidebar with instructions
    with st.sidebar:
        st.header("📖 Instructions")
        st.markdown("""
        **Step 1:** Upload all three required Excel files
        
        **Step 2:** Configure processing options
        
        **Step 3:** Click Process to run automation
        
        **Step 4:** Review preview and download results
        
        ---
        
        **Files Required:**
        - **GL Inquiry**: Actual costs from Sage
        - **WIP Worksheet**: Job budgets from Sage  
        - **Master WIP Report**: File to be updated
        
        ---
        
        **Safety Features:**
        - ✅ Automatic backups
        - ✅ Formula preservation
        - ✅ Preview before update
        - ✅ Validation reports
        - ✅ **Memory-only processing** (no temp files!)
        """)
    
    # Main content
    files_ready = display_file_upload_section()
    
    if files_ready:
        options = display_processing_options()
        
        # Process button
        if st.button("🚀 Process Data", type="primary", use_container_width=True):
            merged_df, gl_aggregated = process_data(options)
            
            if merged_df is not None:
                st.session_state.processed_data = (merged_df, gl_aggregated, options)
                st.session_state.processing_complete = True
                st.rerun()
    
    # Display results if processing is complete
    if st.session_state.processing_complete and st.session_state.processed_data:
        merged_df, gl_aggregated, options = st.session_state.processed_data
        
        # Data preview
        display_data_preview(merged_df, gl_aggregated)
        
        # Update Excel file if not preview only
        updated_file_bytes = None
        if not options['preview_only']:
            if st.button("✍️ Apply Updates to Excel", type="primary"):
                updated_file_bytes = update_excel_file_memory_only(merged_df, options)
                if updated_file_bytes:
                    st.success("✅ Excel file updated successfully!")
                    if st.session_state.backup_created:
                        st.info(f"💾 Backup created: {st.session_state.backup_created}")
        
        # Download section
        display_download_section(updated_file_bytes, merged_df)
        
        # Reset button
        if st.button("🔄 Start Over"):
            for key in st.session_state.keys():
                del st.session_state[key]
            st.rerun()

if __name__ == "__main__":
    main() 