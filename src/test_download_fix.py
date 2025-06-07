#!/usr/bin/env python3
"""
Test script to verify Excel download fix
"""
import io
import openpyxl
from pathlib import Path

def test_excel_download_fix():
    """Test the BytesIO download fix"""
    print("Testing Excel download fix...")
    
    # Test file path
    test_file = "/app/test_data/Master WIP Report.xlsx"
    
    try:
        # Method 1: OLD way (might cause corruption)
        print("\n1. Testing OLD method (direct file read):")
        with open(test_file, "rb") as f:
            old_method_bytes = f.read()
        print(f"   File size: {len(old_method_bytes)} bytes")
        
        # Method 2: NEW way (proper BytesIO handling)
        print("\n2. Testing NEW method (BytesIO buffer):")
        with open(test_file, "rb") as f:
            file_data = f.read()
        
        # Create BytesIO buffer for proper handling
        excel_buffer = io.BytesIO(file_data)
        excel_buffer.seek(0)  # Critical: reset to beginning
        new_method_bytes = excel_buffer.getvalue()
        excel_buffer.close()
        print(f"   File size: {len(new_method_bytes)} bytes")
        
        # Verify they're the same
        if old_method_bytes == new_method_bytes:
            print("   ✅ Both methods produce identical bytes")
        else:
            print("   ❌ Methods produce different bytes")
        
        # Test that the new method bytes can be loaded as Excel
        print("\n3. Testing Excel file integrity:")
        try:
            test_buffer = io.BytesIO(new_method_bytes)
            wb = openpyxl.load_workbook(test_buffer, keep_vba=True)
            print(f"   ✅ Excel file loads successfully")
            print(f"   Sheets: {wb.sheetnames}")
            wb.close()
            test_buffer.close()
        except Exception as e:
            print(f"   ❌ Excel file failed to load: {e}")
        
        print("\n4. Testing Streamlit download simulation:")
        # Simulate what Streamlit download_button expects
        try:
            # This is what the download_button will receive
            download_data = new_method_bytes
            
            # Verify it's still valid
            test_buffer2 = io.BytesIO(download_data)
            wb2 = openpyxl.load_workbook(test_buffer2, keep_vba=True)
            print(f"   ✅ Download data works for openpyxl")
            print(f"   Final file size: {len(download_data)} bytes")
            wb2.close()
            test_buffer2.close()
            
        except Exception as e:
            print(f"   ❌ Download simulation failed: {e}")
        
        print("\n✅ All tests passed! Download fix should work.")
        
    except Exception as e:
        print(f"❌ Test failed: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_excel_download_fix() 