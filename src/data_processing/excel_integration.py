"""
Excel Integration Module

This module handles reading from and writing to the master WIP Report Excel file.
It preserves formulas, formatting, VBA macros, and other Excel features while
updating only the data values in specific sections.
"""

import logging
import os
import shutil
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell


def load_wip_workbook(file_path: str, keep_vba: bool = True) -> Workbook:
    """
    Load the WIP Report workbook with VBA preservation.
    
    Args:
        file_path (str): Path to the WIP Report Excel file
        keep_vba (bool): Whether to preserve VBA macros (default: True)
        
    Returns:
        Workbook: Loaded openpyxl workbook object
        
    Raises:
        FileNotFoundError: If the file doesn't exist
        Exception: If the file cannot be loaded
    """
    try:
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"WIP Report file not found: {file_path}")
        
        # Load workbook with VBA preservation
        workbook = load_workbook(file_path, keep_vba=keep_vba, data_only=False)
        
        logging.info(f"Successfully loaded WIP Report workbook: {file_path}")
        logging.info(f"Worksheets available: {workbook.sheetnames}")
        
        return workbook
        
    except Exception as e:
        logging.error(f"Error loading WIP Report workbook: {str(e)}")
        raise


def find_or_create_monthly_tab(workbook: Workbook, month_year: str) -> Worksheet:
    """
    Find existing monthly tab or create a new one with the specified month/year.
    
    Args:
        workbook (Workbook): The WIP Report workbook
        month_year (str): Month/year in format "MMM YY" (e.g., "Jan 24")
        
    Returns:
        Worksheet: The monthly worksheet
    """
    # Check if the tab already exists
    if month_year in workbook.sheetnames:
        worksheet = workbook[month_year]
        logging.info(f"Found existing monthly tab: {month_year}")
        return worksheet
    
    # Create new tab by copying template (if exists) or creating blank
    template_names = ['Template', 'TEMPLATE', 'template', 'Master']
    template_sheet = None
    
    for template_name in template_names:
        if template_name in workbook.sheetnames:
            template_sheet = workbook[template_name]
            break
    
    if template_sheet:
        # Copy template sheet
        new_sheet = workbook.copy_worksheet(template_sheet)
        new_sheet.title = month_year
        logging.info(f"Created new monthly tab '{month_year}' from template '{template_sheet.title}'")
    else:
        # Create blank sheet
        new_sheet = workbook.create_sheet(title=month_year)
        logging.info(f"Created new blank monthly tab: {month_year}")
    
    return new_sheet


def find_section_markers(worksheet: Worksheet, markers: List[str]) -> Dict[str, Optional[Tuple[int, int]]]:
    """
    Find the start positions of sections marked by specific strings.
    
    Args:
        worksheet (Worksheet): The worksheet to search
        markers (List[str]): List of marker strings to find (e.g., ['5040', '5030'])
        
    Returns:
        Dict[str, Optional[Tuple[int, int]]]: Dictionary mapping markers to (row, col) positions
    """
    section_positions = {}
    
    for marker in markers:
        section_positions[marker] = None
        
        # Search through all cells for the marker
        for row in range(1, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                if cell.value and marker in str(cell.value).strip():
                    section_positions[marker] = (row, col)
                    logging.info(f"Found section marker '{marker}' at row {row}, column {col} (cell value: '{cell.value}')")
                    break
            if section_positions[marker]:
                break
        
        if not section_positions[marker]:
            logging.warning(f"Section marker '{marker}' not found in worksheet")
    
    return section_positions


def detect_data_region(worksheet: Worksheet, start_row: int, start_col: int, 
                      max_rows: int = 200) -> Tuple[int, int]:
    """
    Detect the data region starting from a given position until empty rows are found.
    
    Args:
        worksheet (Worksheet): The worksheet to analyze
        start_row (int): Starting row number
        start_col (int): Starting column number
        max_rows (int): Maximum number of rows to scan
        
    Returns:
        Tuple[int, int]: (end_row, end_col) of the data region
    """
    end_row = start_row
    end_col = start_col
    
    # Find the end of data by looking for empty rows
    consecutive_empty_rows = 0
    for row in range(start_row, start_row + max_rows):
        row_has_data = False
        
        # Check if this row has any data
        for col in range(start_col, start_col + 20):  # Check up to 20 columns
            cell = worksheet.cell(row=row, column=col)
            if cell.value is not None and str(cell.value).strip():
                row_has_data = True
                end_col = max(end_col, col)
                break
        
        if row_has_data:
            end_row = row
            consecutive_empty_rows = 0
        else:
            consecutive_empty_rows += 1
            # Stop if we find 3 consecutive empty rows
            if consecutive_empty_rows >= 3:
                break
    
    logging.info(f"Detected data region from ({start_row}, {start_col}) to ({end_row}, {end_col})")
    return end_row, end_col


def is_formula_cell(cell: Cell) -> bool:
    """
    Check if a cell contains a formula.
    
    Args:
        cell (Cell): The cell to check
        
    Returns:
        bool: True if the cell contains a formula
    """
    return cell.data_type == 'f' or (cell.value and str(cell.value).startswith('='))


def clear_data_preserve_formulas(worksheet: Worksheet, start_row: int, end_row: int, 
                                start_col: int, end_col: int) -> int:
    """
    Clear only data values in the specified range, preserving formulas and formatting.
    
    Args:
        worksheet (Worksheet): The worksheet to modify
        start_row (int): Starting row number
        end_row (int): Ending row number
        start_col (int): Starting column number
        end_col (int): Ending column number
        
    Returns:
        int: Number of cells cleared
    """
    cells_cleared = 0
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = worksheet.cell(row=row, column=col)
            
            # Only clear if it's not a formula cell
            if not is_formula_cell(cell) and cell.value is not None:
                cell.value = None
                cells_cleared += 1
    
    logging.info(f"Cleared {cells_cleared} data cells while preserving formulas")
    return cells_cleared


def write_job_data_to_section(worksheet: Worksheet, job_data: pd.DataFrame, 
                             section_start_row: int, section_start_col: int,
                             job_col_offset: int = 0, data_col_offset: int = 1) -> int:
    """
    Write job data to a specific section of the worksheet.
    
    Args:
        worksheet (Worksheet): The worksheet to write to
        job_data (pd.DataFrame): DataFrame containing job data
        section_start_row (int): Starting row of the section
        section_start_col (int): Starting column of the section
        job_col_offset (int): Column offset for job numbers (default: 0)
        data_col_offset (int): Column offset for data values (default: 1)
        
    Returns:
        int: Number of jobs written
    """
    jobs_written = 0
    
    # Determine which data column to write based on the DataFrame columns
    data_column = None
    if 'Material' in job_data.columns:
        data_column = 'Material'
    elif 'Labor' in job_data.columns:
        data_column = 'Labor'
    elif 'Other' in job_data.columns:
        data_column = 'Other'
    else:
        logging.warning("No recognized data column found in job_data")
        return 0
    
    # Write each job's data
    for idx, (_, row) in enumerate(job_data.iterrows()):
        current_row = section_start_row + 1 + idx  # +1 to skip header row
        
        # Write job number
        job_cell = worksheet.cell(row=current_row, column=section_start_col + job_col_offset)
        if not is_formula_cell(job_cell):
            job_cell.value = row['Job Number']
        
        # Write data value
        data_cell = worksheet.cell(row=current_row, column=section_start_col + data_col_offset)
        if not is_formula_cell(data_cell):
            data_cell.value = row[data_column]
            jobs_written += 1
    
    logging.info(f"Wrote data for {jobs_written} jobs to section starting at row {section_start_row}")
    return jobs_written


def create_backup(file_path: str, backup_dir: str = "WIP_Backups") -> str:
    """
    Create a timestamped backup of the WIP Report file.
    
    Args:
        file_path (str): Path to the original file
        backup_dir (str): Directory to store backups (default: "WIP_Backups")
        
    Returns:
        str: Path to the backup file
        
    Raises:
        Exception: If backup creation fails
    """
    try:
        # Create backup directory if it doesn't exist
        backup_path = Path(backup_dir)
        backup_path.mkdir(exist_ok=True)
        
        # Generate timestamped backup filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        original_name = Path(file_path).stem
        original_ext = Path(file_path).suffix
        backup_filename = f"{original_name}_BACKUP_{timestamp}{original_ext}"
        backup_file_path = backup_path / backup_filename
        
        # Copy the file
        shutil.copy2(file_path, backup_file_path)
        
        logging.info(f"Created backup: {backup_file_path}")
        return str(backup_file_path)
        
    except Exception as e:
        logging.error(f"Failed to create backup: {str(e)}")
        raise


def update_wip_report(file_path: str, material_data: pd.DataFrame, labor_data: pd.DataFrame,
                     month_year: str, create_backup_flag: bool = True) -> Dict[str, Any]:
    """
    Update the WIP Report with new material and labor data.
    
    Args:
        file_path (str): Path to the WIP Report Excel file
        material_data (pd.DataFrame): Material data for 5040 section
        labor_data (pd.DataFrame): Labor data for 5030 section
        month_year (str): Month/year for the tab (e.g., "Jan 24")
        create_backup_flag (bool): Whether to create a backup before updating
        
    Returns:
        Dict[str, Any]: Summary of the update operation
    """
    summary = {
        'backup_created': None,
        'jobs_updated': {'material': 0, 'labor': 0},
        'cells_cleared': 0,
        'success': False,
        'error': None
    }
    
    try:
        # Create backup if requested
        if create_backup_flag:
            backup_path = create_backup(file_path)
            summary['backup_created'] = backup_path
        
        # Load the workbook
        workbook = load_wip_workbook(file_path)
        
        # Find or create the monthly tab
        worksheet = find_or_create_monthly_tab(workbook, month_year)
        
        # Find section markers
        section_positions = find_section_markers(worksheet, ['5040', '5030'])
        
        # Update Material section (5040)
        if section_positions['5040'] and not material_data.empty:
            start_row, start_col = section_positions['5040']
            end_row, end_col = detect_data_region(worksheet, start_row + 1, start_col)
            
            # Clear existing data
            cells_cleared = clear_data_preserve_formulas(worksheet, start_row + 1, end_row, 
                                                       start_col, end_col)
            summary['cells_cleared'] += cells_cleared
            
            # Write new material data
            jobs_written = write_job_data_to_section(worksheet, material_data, start_row, start_col)
            summary['jobs_updated']['material'] = jobs_written
        
        # Update Labor section (5030)
        if section_positions['5030'] and not labor_data.empty:
            start_row, start_col = section_positions['5030']
            end_row, end_col = detect_data_region(worksheet, start_row + 1, start_col)
            
            # Clear existing data
            cells_cleared = clear_data_preserve_formulas(worksheet, start_row + 1, end_row, 
                                                       start_col, end_col)
            summary['cells_cleared'] += cells_cleared
            
            # Write new labor data
            jobs_written = write_job_data_to_section(worksheet, labor_data, start_row, start_col)
            summary['jobs_updated']['labor'] = jobs_written
        
        # Save the workbook
        workbook.save(file_path)
        summary['success'] = True
        
        logging.info(f"Successfully updated WIP Report: {summary}")
        
    except Exception as e:
        error_msg = f"Error updating WIP Report: {str(e)}"
        logging.error(error_msg)
        summary['error'] = error_msg
        
        # If we created a backup and there was an error, we could restore it here
        if summary['backup_created'] and create_backup_flag:
            try:
                shutil.copy2(summary['backup_created'], file_path)
                logging.info(f"Restored original file from backup due to error")
            except Exception as restore_error:
                logging.error(f"Failed to restore backup: {str(restore_error)}")
    
    return summary


def get_existing_data_from_section(worksheet: Worksheet, section_marker: str) -> pd.DataFrame:
    """
    Extract existing data from a section for comparison purposes.
    
    Args:
        worksheet (Worksheet): The worksheet to read from
        section_marker (str): Section marker to find (e.g., '5040', '5030')
        
    Returns:
        pd.DataFrame: Existing data from the section
    """
    section_positions = find_section_markers(worksheet, [section_marker])
    
    if not section_positions[section_marker]:
        logging.warning(f"Section marker '{section_marker}' not found")
        return pd.DataFrame()
    
    start_row, start_col = section_positions[section_marker]
    end_row, end_col = detect_data_region(worksheet, start_row + 1, start_col)
    
    data = []
    for row in range(start_row + 1, end_row + 1):
        job_cell = worksheet.cell(row=row, column=start_col)
        data_cell = worksheet.cell(row=row, column=start_col + 1)
        
        if job_cell.value:
            data.append({
                'Job Number': str(job_cell.value).strip(),
                'Current Value': data_cell.value if data_cell.value is not None else 0
            })
    
    df = pd.DataFrame(data)
    logging.info(f"Extracted {len(df)} existing records from section '{section_marker}'")
    return df


if __name__ == "__main__":
    # Example usage and testing
    logging.basicConfig(level=logging.INFO)
    
    # This would be used for testing with a sample file
    # sample_material_data = pd.DataFrame({
    #     'Job Number': ['JOB001', 'JOB002'],
    #     'Material': [10000, 5000]
    # })
    # 
    # sample_labor_data = pd.DataFrame({
    #     'Job Number': ['JOB001', 'JOB002'],
    #     'Labor': [8000, 4000]
    # })
    # 
    # result = update_wip_report('WIP_Report.xlsx', sample_material_data, sample_labor_data, 'Jan 24')
    # print(result)
    pass 