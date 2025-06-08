"""
Surgical Excel Update Module

This module implements a surgical approach to Excel file modification that:
1. Treats Excel files as ZIP archives (which they are)
2. Surgically modifies only specific cell values in the XML
3. Preserves ALL other content: formulas, formatting, VBA, metadata
4. Avoids the 56KB data loss issue with openpyxl

Based on Claude's enterprise-grade approach for critical financial files.
"""

import zipfile
import xml.etree.ElementTree as ET
import io
import re
import logging
from typing import Dict, Any, Optional, Tuple
from openpyxl import load_workbook

# Set up logging
logger = logging.getLogger(__name__)

def find_sheet_id_from_workbook(file_bytes: bytes, sheet_name: str) -> Optional[str]:
    """
    Use openpyxl ONLY to find the sheet ID - no modifications
    """
    try:
        # Use openpyxl in read-only mode just to get sheet structure
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        
        # Find the sheet
        if sheet_name not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
            wb.close()
            return None
        
        # Get sheet index (1-based for XML files)
        sheet_index = wb.sheetnames.index(sheet_name) + 1
        wb.close()
        
        return str(sheet_index)
        
    except Exception as e:
        logger.error(f"Error finding sheet ID: {e}")
        return None

def find_section_locations(file_bytes: bytes, sheet_name: str) -> Tuple[Optional[int], Optional[int]]:
    """
    Use openpyxl ONLY to find section locations - no modifications
    Returns (row_5040, row_5030) or (None, None) if not found
    """
    try:
        # Use openpyxl in read-only mode just to scan for sections
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        
        if sheet_name not in wb.sheetnames:
            logger.error(f"Sheet '{sheet_name}' not found")
            wb.close()
            return None, None
        
        ws = wb[sheet_name]
        
        row_5040 = None
        row_5030 = None
        
        # Scan for section markers
        for row in ws.iter_rows(max_col=10, max_row=200):  # Reasonable search range
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    cell_text = str(cell.value).lower().strip()
                    
                    if '5040' in cell_text and 'labor' in cell_text:
                        row_5040 = cell.row
                        logger.info(f"Found 5040 section at row {row_5040}")
                    
                    if '5030' in cell_text and 'material' in cell_text:
                        row_5030 = cell.row
                        logger.info(f"Found 5030 section at row {row_5030}")
        
        wb.close()
        return row_5040, row_5030
        
    except Exception as e:
        logger.error(f"Error finding sections: {e}")
        return None, None

def column_letter_to_number(column_letter: str) -> int:
    """Convert Excel column letter to number (A=1, B=2, etc.)"""
    result = 0
    for char in column_letter.upper():
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result

def number_to_column_letter(number: int) -> str:
    """Convert number to Excel column letter (1=A, 2=B, etc.)"""
    result = ""
    while number > 0:
        number -= 1
        result = chr(number % 26 + ord('A')) + result
        number //= 26
    return result

def update_cell_in_xml(root: ET.Element, cell_ref: str, value: Any) -> bool:
    """
    Update a specific cell value in the worksheet XML
    Returns True if cell was found and updated, False otherwise
    """
    try:
        # Define namespace
        ns = {'': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
        
        # Find the cell
        for row_elem in root.findall('.//row', ns):
            for cell_elem in row_elem.findall('c', ns):
                if cell_elem.get('r') == cell_ref:
                    # Found the cell - update its value
                    v_elem = cell_elem.find('v', ns)
                    if v_elem is None:
                        # Create value element if it doesn't exist
                        v_elem = ET.SubElement(cell_elem, 'v')
                    
                    # Set the value
                    v_elem.text = str(value) if value is not None else ""
                    
                    # Set appropriate type
                    if isinstance(value, (int, float)) and value != "":
                        cell_elem.set('t', 'n')  # Number
                    else:
                        # Remove type attribute for text (default)
                        if 't' in cell_elem.attrib:
                            del cell_elem.attrib['t']
                    
                    logger.debug(f"Updated cell {cell_ref} = {value}")
                    return True
        
        logger.warning(f"Cell {cell_ref} not found in XML")
        return False
        
    except Exception as e:
        logger.error(f"Error updating cell {cell_ref}: {e}")
        return False

def surgical_excel_update(excel_bytes: bytes, cell_updates: Dict[str, Dict[str, Any]]) -> bytes:
    """
    Surgically update specific cells in Excel without corrupting the file
    
    Args:
        excel_bytes: Original Excel file as bytes
        cell_updates: Dictionary like:
            {
                'Apr 25': {  # sheet name
                    'A6': 'New Value',
                    'B6': 'Another Value', 
                    'E6': 12345.67
                }
            }
    
    Returns:
        Updated Excel file as bytes
    """
    try:
        logger.info(f"Starting surgical update on {len(excel_bytes)} byte file")
        logger.info(f"Updates for sheets: {list(cell_updates.keys())}")
        
        # Create in-memory zip file from the Excel bytes
        excel_zip = io.BytesIO(excel_bytes)
        
        # Read all files from the zip
        file_dict = {}
        with zipfile.ZipFile(excel_zip, 'r') as zip_in:
            file_dict = {name: zip_in.read(name) for name in zip_in.namelist()}
        
        logger.info(f"Read {len(file_dict)} files from Excel ZIP archive")
        
        # Process each sheet that needs updates
        for sheet_name, updates in cell_updates.items():
            if not updates:
                continue
                
            logger.info(f"Processing {len(updates)} updates for sheet '{sheet_name}'")
            
            # Find the sheet ID
            sheet_id = find_sheet_id_from_workbook(excel_bytes, sheet_name)
            if not sheet_id:
                logger.error(f"Could not find sheet ID for '{sheet_name}'")
                continue
            
            # Construct the worksheet path
            sheet_path = f'xl/worksheets/sheet{sheet_id}.xml'
            
            if sheet_path not in file_dict:
                logger.error(f"Worksheet file {sheet_path} not found in archive")
                continue
            
            # Parse the worksheet XML
            try:
                root = ET.fromstring(file_dict[sheet_path])
            except ET.ParseError as e:
                logger.error(f"Error parsing XML for {sheet_path}: {e}")
                continue
            
            # Update each cell
            updated_count = 0
            for cell_ref, new_value in updates.items():
                if update_cell_in_xml(root, cell_ref, new_value):
                    updated_count += 1
            
            logger.info(f"Successfully updated {updated_count}/{len(updates)} cells in {sheet_name}")
            
            # Convert updated XML back to bytes
            file_dict[sheet_path] = ET.tostring(root, encoding='utf-8', xml_declaration=True)
        
        # Create new ZIP with all files (modified and unmodified)
        output = io.BytesIO()
        with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED, compresslevel=6) as zip_out:
            for filename, content in file_dict.items():
                zip_out.writestr(filename, content)
        
        output.seek(0)
        result_bytes = output.getvalue()
        
        logger.info(f"Surgical update complete. Output size: {len(result_bytes)} bytes")
        logger.info(f"Size change: {len(result_bytes) - len(excel_bytes):+d} bytes")
        
        return result_bytes
        
    except Exception as e:
        logger.error(f"Surgical update failed: {e}")
        # Return original bytes on failure
        return excel_bytes

def update_wip_report_surgical(
    master_file_bytes: bytes, 
    merged_df, 
    month_year: str
) -> bytes:
    """
    Update WIP Report using surgical approach
    
    This function:
    1. Finds the section locations using openpyxl (read-only)
    2. Maps the data to specific cell updates
    3. Uses surgical ZIP/XML modification to preserve everything
    """
    try:
        logger.info(f"Starting surgical WIP report update for {month_year}")
        
        # Find section locations
        row_5040, row_5030 = find_section_locations(master_file_bytes, month_year)
        
        if not row_5040 and not row_5030:
            logger.error("Could not find any section markers")
            return master_file_bytes
        
        # Prepare cell updates
        cell_updates = {month_year: {}}
        
        # Update 5040 section (Sub Labor)
        if row_5040 and 'Sub Labor' in merged_df.columns:
            current_row = row_5040 + 1  # Start one row below header
            
            for _, job_data in merged_df.iterrows():
                job_number = str(job_data.get('Job Number', '')).strip()
                sub_labor = job_data.get('Sub Labor', 0)
                
                if job_number and sub_labor != 0:
                    # Job number in column A, Sub Labor in column E (typical WIP layout)
                    cell_updates[month_year][f'A{current_row}'] = job_number
                    cell_updates[month_year][f'E{current_row}'] = float(sub_labor) if sub_labor else 0
                    current_row += 1
        
        # Update 5030 section (Material)  
        if row_5030 and 'Material' in merged_df.columns:
            current_row = row_5030 + 1  # Start one row below header
            
            for _, job_data in merged_df.iterrows():
                job_number = str(job_data.get('Job Number', '')).strip()
                material = job_data.get('Material', 0)
                
                if job_number and material != 0:
                    # Job number in column A, Material in column E
                    cell_updates[month_year][f'A{current_row}'] = job_number
                    cell_updates[month_year][f'E{current_row}'] = float(material) if material else 0
                    current_row += 1
        
        logger.info(f"Prepared {len(cell_updates[month_year])} cell updates")
        
        # Perform surgical update
        return surgical_excel_update(master_file_bytes, cell_updates)
        
    except Exception as e:
        logger.error(f"Error in surgical WIP update: {e}")
        return master_file_bytes

def create_backup_from_bytes(file_bytes: bytes, backup_folder: str = "WIP_Backups") -> str:
    """
    Create a backup file from bytes
    Returns the backup filename
    """
    import os
    from datetime import datetime
    
    try:
        # Create backup folder if it doesn't exist
        os.makedirs(backup_folder, exist_ok=True)
        
        # Generate backup filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"WIP_Report_BACKUP_{timestamp}.xlsx"
        backup_path = os.path.join(backup_folder, backup_filename)
        
        # Write backup file
        with open(backup_path, 'wb') as f:
            f.write(file_bytes)
        
        logger.info(f"Created backup: {backup_path}")
        return backup_filename
        
    except Exception as e:
        logger.error(f"Error creating backup: {e}")
        return "" 