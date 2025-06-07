#!/usr/bin/env python3
"""
Enhanced Excel Integration Module for WIP Report Automation
Handles complex column structures for 5040 (Sub Labor) and 5030 (Material) sections
"""

import pandas as pd
import logging
from pathlib import Path
from datetime import datetime
import shutil
from typing import Dict, List, Optional, Tuple, Any
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell


def load_wip_workbook(file_path: str, keep_vba: bool = True) -> Workbook:
    """
    Load the WIP Report workbook with VBA preservation.
    
    Args:
        file_path (str): Path to the WIP Report Excel file
        keep_vba (bool): Whether to preserve VBA macros (default: True)
        
    Returns:
        Workbook: The loaded openpyxl workbook object
        
    Raises:
        FileNotFoundError: If the file doesn't exist
        Exception: If the workbook cannot be loaded
    """
    try:
        if not Path(file_path).exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        workbook = load_workbook(file_path, keep_vba=keep_vba)
        logging.info(f"Successfully loaded WIP workbook: {file_path}")
        return workbook
        
    except Exception as e:
        logging.error(f"Error loading WIP workbook: {str(e)}")
        raise


def find_or_create_monthly_tab(workbook: Workbook, month_year: str) -> Worksheet:
    """
    Find existing monthly tab or create new one based on month/year.
    Always standardizes to 3-letter month + 2-digit year format (e.g., "Apr 25").
    
    Args:
        workbook (Workbook): The WIP Report workbook
        month_year (str): Month/year string (e.g., "Jun 25", "June 25", "April 2025")
        
    Returns:
        Worksheet: The monthly worksheet
    """
    # Standardize to 3-letter format for all operations
    month_map = {
        'january': 'Jan', 'february': 'Feb', 'march': 'Mar', 'april': 'Apr',
        'may': 'May', 'june': 'Jun', 'july': 'Jul', 'august': 'Aug',
        'september': 'Sep', 'october': 'Oct', 'november': 'Nov', 'december': 'Dec'
    }
    
    # Convert input to standardized format
    standard_name = month_year
    
    # Handle full year format (e.g., "April 2025" -> "Apr 25")
    for full_month, short_month in month_map.items():
        if full_month.lower() in month_year.lower():
            # Extract year (last 2 digits)
            year_match = month_year.split()[-1] if ' ' in month_year else '25'
            if len(year_match) == 4:  # Full year like "2025"
                year_short = year_match[-2:]  # Get last 2 digits
            else:
                year_short = year_match
            standard_name = f"{short_month} {year_short}"
            break
    
    # Look for existing tab (check both standard and possible variations)
    target_month = standard_name.split()[0]  # e.g., "Apr"
    target_year = standard_name.split()[1]   # e.g., "25"
    
    for sheet_name in workbook.sheetnames:
        sheet_lower = sheet_name.lower().strip()
        
        # Check for exact match first
        if standard_name.lower() == sheet_lower:
            logging.info(f"Found existing monthly tab (exact): {sheet_name}")
            return workbook[sheet_name]
        
        # Check for month/year match with variations
        if (target_month.lower() in sheet_lower and 
            target_year in sheet_lower):
            logging.info(f"Found existing monthly tab (variation): {sheet_name}")
            return workbook[sheet_name]
    
    # Create new tab with standardized name
    new_worksheet = workbook.create_sheet(title=standard_name)
    logging.info(f"Created new monthly tab: {standard_name}")
    return new_worksheet


def find_section_markers(worksheet: Worksheet, section_patterns: List[str]) -> Dict[str, Optional[Tuple[int, int]]]:
    """
    Find section markers in the worksheet using pattern matching.
    
    Args:
        worksheet (Worksheet): The worksheet to search
        section_patterns (List[str]): Patterns to search for (e.g., ['5040', '5030'])
        
    Returns:
        Dict[str, Optional[Tuple[int, int]]]: Section positions {pattern: (row, col)}
    """
    section_positions = {}
    
    # Extended patterns to match the real headers
    pattern_mappings = {
        '5040': ['5040', '% of sub labor cost', 'sub labor cost - 5040', '% of sub labor cost - 5040'],
        '5030': ['5030', '% of material', 'material - 5030', '% of material - 5030']
    }
    
    for pattern in section_patterns:
        section_positions[pattern] = None
        search_patterns = pattern_mappings.get(pattern, [pattern])
        
        # Search through all cells for the markers
        for row in range(1, min(worksheet.max_row + 1, 100)):  # Search first 100 rows
            for col in range(1, min(worksheet.max_column + 1, 10)):  # Search first 10 columns
                cell = worksheet.cell(row=row, column=col)
                if cell.value:
                    cell_text = str(cell.value).lower().strip()
                    
                    # Check if any pattern matches
                    for search_pattern in search_patterns:
                        if search_pattern.lower() in cell_text:
                            section_positions[pattern] = (row, col)
                            logging.info(f"Found section marker '{pattern}' at row {row}, column {col}: '{cell.value}'")
                            break
                    
                    if section_positions[pattern]:
                        break
            
            if section_positions[pattern]:
                break
        
        if not section_positions[pattern]:
            logging.warning(f"Section marker '{pattern}' not found in worksheet")
    
    return section_positions


def is_formula_cell(cell: Cell) -> bool:
    """
    Check if a cell contains a formula.
    
    Args:
        cell (Cell): The cell to check
        
    Returns:
        bool: True if the cell contains a formula
    """
    return cell.data_type == 'f' or (cell.value and str(cell.value).startswith('='))


def is_merged_cell(worksheet: Worksheet, row: int, col: int) -> bool:
    """
    Check if a cell is part of a merged range.
    
    Args:
        worksheet (Worksheet): The worksheet to check
        row (int): Row number
        col (int): Column number
        
    Returns:
        bool: True if the cell is merged
    """
    cell_coordinate = worksheet.cell(row=row, column=col).coordinate
    
    for merged_range in worksheet.merged_cells.ranges:
        if cell_coordinate in merged_range:
            return True
    return False


def get_merged_cell_top_left(worksheet: Worksheet, row: int, col: int) -> tuple:
    """
    If a cell is part of a merged range, return the top-left cell coordinates.
    Otherwise, return the original coordinates.
    
    Args:
        worksheet (Worksheet): The worksheet to check
        row (int): Row number
        col (int): Column number
        
    Returns:
        tuple: (row, col) of the top-left cell of merged range, or original if not merged
    """
    cell_coordinate = worksheet.cell(row=row, column=col).coordinate
    
    for merged_range in worksheet.merged_cells.ranges:
        if cell_coordinate in merged_range:
            # Return top-left cell of the merged range
            return merged_range.min_row, merged_range.min_col
    
    # Not merged, return original coordinates
    return row, col


def safe_write_cell(worksheet: Worksheet, row: int, col: int, value) -> bool:
    """
    Safely write to a cell, handling merged cells by writing to the top-left cell.
    
    Args:
        worksheet (Worksheet): The worksheet to write to
        row (int): Row number
        col (int): Column number
        value: Value to write
        
    Returns:
        bool: True if successfully written, False if skipped
    """
    try:
        # If it's a merged cell, get the top-left coordinates
        actual_row, actual_col = get_merged_cell_top_left(worksheet, row, col)
        
        if (actual_row, actual_col) != (row, col):
            logging.info(f"Writing to merged cell: redirecting from ({row},{col}) to top-left ({actual_row},{actual_col})")
        
        cell = worksheet.cell(row=actual_row, column=actual_col)
        
        # Skip if it's a formula
        if is_formula_cell(cell):
            logging.warning(f"Skipping formula cell at row {actual_row}, col {actual_col}")
            return False
        
        # Write the value
        cell.value = value
        return True
        
    except Exception as e:
        # If ANY error occurs during writing, log it and skip
        error_msg = str(e)
        if "MergedCell" in error_msg or "read-only" in error_msg:
            logging.warning(f"Skipping problematic cell at row {row}, col {col}: {error_msg}")
        else:
            logging.error(f"Error writing to cell {row},{col}: {error_msg}")
        return False


def clear_data_preserve_formulas_5040(worksheet: Worksheet, start_row: int, 
                                     job_col: int = 1, data_cols: List[int] = [3, 4, 5, 8],
                                     max_rows: int = 200) -> int:
    """
    Clear data in 5040 section while preserving formulas.
    
    Args:
        worksheet (Worksheet): The worksheet to modify
        start_row (int): Starting row (section header row)
        job_col (int): Column for job numbers (default: 1 = Column A)
        data_cols (List[int]): Columns with data to clear (default: [3,4,5,8] = C,D,E,H)
        max_rows (int): Maximum rows to process
        
    Returns:
        int: Number of cells cleared
    """
    cells_cleared = 0
    
    # Start clearing from the row after the header
    current_row = start_row + 1
    consecutive_empty_rows = 0
    
    for row_offset in range(max_rows):
        row_num = current_row + row_offset
        row_has_data = False
        
        # Check if this row has job data
        job_cell = worksheet.cell(row=row_num, column=job_col)
        if job_cell.value and str(job_cell.value).strip():
            row_has_data = True
            
            # Clear job number (safe_write_cell handles merged cells)
            if safe_write_cell(worksheet, row_num, job_col, None):
                cells_cleared += 1
            
            # Clear data columns (safe_write_cell handles merged cells)
            for col_num in data_cols:
                if safe_write_cell(worksheet, row_num, col_num, None):
                    cells_cleared += 1
        
        if row_has_data:
            consecutive_empty_rows = 0
        else:
            consecutive_empty_rows += 1
            # Stop if we find 3 consecutive empty rows
            if consecutive_empty_rows >= 3:
                break
    
    logging.info(f"Cleared {cells_cleared} cells in 5040 section starting at row {start_row}")
    return cells_cleared


def clear_data_preserve_formulas_5030(worksheet: Worksheet, start_row: int,
                                     desc_col: int = 1, data_cols: List[int] = [2, 3],
                                     max_rows: int = 200) -> int:
    """
    Clear data in 5030 section while preserving formulas.
    
    Args:
        worksheet (Worksheet): The worksheet to modify
        start_row (int): Starting row (section header row)
        desc_col (int): Column for job descriptions (default: 1 = Column A)
        data_cols (List[int]): Columns with data to clear (default: [2,3] = B,C)
        max_rows (int): Maximum rows to process
        
    Returns:
        int: Number of cells cleared
    """
    cells_cleared = 0
    
    # Start clearing from the row after the header
    current_row = start_row + 1
    consecutive_empty_rows = 0
    
    for row_offset in range(max_rows):
        row_num = current_row + row_offset
        row_has_data = False
        
        # Check if this row has description data
        desc_cell = worksheet.cell(row=row_num, column=desc_col)
        if desc_cell.value and str(desc_cell.value).strip():
            row_has_data = True
            
            # Clear description (safe_write_cell handles merged cells)
            if safe_write_cell(worksheet, row_num, desc_col, None):
                cells_cleared += 1
            
            # Clear data columns (safe_write_cell handles merged cells)
            for col_num in data_cols:
                if safe_write_cell(worksheet, row_num, col_num, None):
                    cells_cleared += 1
        
        if row_has_data:
            consecutive_empty_rows = 0
        else:
            consecutive_empty_rows += 1
            # Stop if we find 3 consecutive empty rows
            if consecutive_empty_rows >= 3:
                break
    
    logging.info(f"Cleared {cells_cleared} cells in 5030 section starting at row {start_row}")
    return cells_cleared


def write_5040_section_data(worksheet: Worksheet, job_data: pd.DataFrame, start_row: int,
                           job_col: int = 1, amount_col: int = 5) -> int:
    """
    Write Sub Labor data to the 5040 section.
    
    Args:
        worksheet (Worksheet): The worksheet to write to
        job_data (pd.DataFrame): DataFrame with Sub Labor data
        start_row (int): Starting row of the section
        job_col (int): Column for job numbers (default: 1 = Column A)
        amount_col (int): Column for amounts (default: 5 = Column E)
        
    Returns:
        int: Number of jobs written
    """
    jobs_written = 0
    
    # Filter for jobs with Sub Labor amounts
    sub_labor_jobs = job_data[job_data['Sub Labor'] != 0].copy() if 'Sub Labor' in job_data.columns else pd.DataFrame()
    
    if sub_labor_jobs.empty:
        logging.info("No Sub Labor data to write to 5040 section")
        return 0
    
    # Write each job's data starting from row after header
    for idx, (_, row) in enumerate(sub_labor_jobs.iterrows()):
        current_row = start_row + 1 + idx
        
        # Write job number in Column A
        if safe_write_cell(worksheet, current_row, job_col, row['Job Number']):
            # Write Sub Labor amount in specified column (default Column E)
            if safe_write_cell(worksheet, current_row, amount_col, row['Sub Labor']):
                jobs_written += 1
    
    logging.info(f"Wrote {jobs_written} Sub Labor jobs to 5040 section")
    return jobs_written


def write_5030_section_data(worksheet: Worksheet, job_data: pd.DataFrame, start_row: int,
                           desc_col: int = 1, amount_col: int = 3) -> int:
    """
    Write Material data to the 5030 section.
    
    Args:
        worksheet (Worksheet): The worksheet to write to
        job_data (pd.DataFrame): DataFrame with Material data
        start_row (int): Starting row of the section
        desc_col (int): Column for job descriptions (default: 1 = Column A)
        amount_col (int): Column for amounts (default: 3 = Column C)
        
    Returns:
        int: Number of jobs written
    """
    jobs_written = 0
    
    # Filter for jobs with Material amounts
    material_jobs = job_data[job_data['Material'] != 0].copy() if 'Material' in job_data.columns else pd.DataFrame()
    
    if material_jobs.empty:
        logging.info("No Material data to write to 5030 section")
        return 0
    
    # Write each job's data starting from row after header
    for idx, (_, row) in enumerate(material_jobs.iterrows()):
        current_row = start_row + 1 + idx
        
        # Write job description in Column A
        description = row.get('Job Name', row['Job Number'])
        if safe_write_cell(worksheet, current_row, desc_col, description):
            # Write Material amount in specified column (default Column C)
            if safe_write_cell(worksheet, current_row, amount_col, row['Material']):
                jobs_written += 1
    
    logging.info(f"Wrote {jobs_written} Material jobs to 5030 section")
    return jobs_written


def create_backup(file_path: str, backup_dir: str = "WIP_Backups") -> str:
    """
    Create a timestamped backup of the WIP Report file.
    
    Args:
        file_path (str): Path to the original file
        backup_dir (str): Directory to store backups (default: "WIP_Backups")
        
    Returns:
        str: Path to the backup file
        
    Raises:
        Exception: If backup creation fails
    """
    try:
        # Create backup directory if it doesn't exist
        backup_path = Path(backup_dir)
        backup_path.mkdir(exist_ok=True)
        
        # Generate timestamped backup filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        original_name = Path(file_path).stem
        original_ext = Path(file_path).suffix
        backup_filename = f"{original_name}_BACKUP_{timestamp}{original_ext}"
        backup_file_path = backup_path / backup_filename
        
        # Copy the file
        shutil.copy2(file_path, backup_file_path)
        
        logging.info(f"Created backup: {backup_file_path}")
        return str(backup_file_path)
        
    except Exception as e:
        logging.error(f"Failed to create backup: {str(e)}")
        raise


def update_wip_report_v2(file_path: str, sub_labor_data: pd.DataFrame, material_data: pd.DataFrame,
                        month_year: str, create_backup_flag: bool = True) -> Dict[str, Any]:
    """
    Update the WIP Report with new Sub Labor and Material data using correct column mapping.
    
    Args:
        file_path (str): Path to the WIP Report Excel file
        sub_labor_data (pd.DataFrame): Sub Labor data for 5040 section
        material_data (pd.DataFrame): Material data for 5030 section
        month_year (str): Month/year for the tab (e.g., "Jun 25")
        create_backup_flag (bool): Whether to create a backup before updating
        
    Returns:
        Dict[str, Any]: Summary of the update operation
    """
    summary = {
        'backup_created': None,
        'jobs_updated': {'sub_labor': 0, 'material': 0},
        'cells_cleared': 0,
        'success': False,
        'error': None
    }
    
    try:
        # Create backup if requested
        if create_backup_flag:
            backup_path = create_backup(file_path)
            summary['backup_created'] = backup_path
        
        # Load the workbook
        workbook = load_wip_workbook(file_path)
        
        # Find or create the monthly tab
        worksheet = find_or_create_monthly_tab(workbook, month_year)
        
        # Find section markers
        section_positions = find_section_markers(worksheet, ['5040', '5030'])
        
        # Validate that we found the required sections
        if not section_positions['5040']:
            raise ValueError(f"5040 section not found in worksheet '{worksheet.title}'. Cannot proceed with update.")
        
        if not section_positions['5030']:
            raise ValueError(f"5030 section not found in worksheet '{worksheet.title}'. Cannot proceed with update.")
        
        logging.info(f"Found sections: 5040 at {section_positions['5040']}, 5030 at {section_positions['5030']}")
        
        # Update 5040 section (Sub Labor)
        if not sub_labor_data.empty:
            start_row, start_col = section_positions['5040']
            
            # Clear existing data
            cells_cleared = clear_data_preserve_formulas_5040(worksheet, start_row)
            summary['cells_cleared'] += cells_cleared
            
            # Write new Sub Labor data
            jobs_written = write_5040_section_data(worksheet, sub_labor_data, start_row)
            summary['jobs_updated']['sub_labor'] = jobs_written
        
        # Update 5030 section (Material)
        if not material_data.empty:
            start_row, start_col = section_positions['5030']
            
            # Clear existing data
            cells_cleared = clear_data_preserve_formulas_5030(worksheet, start_row)
            summary['cells_cleared'] += cells_cleared
            
            # Write new Material data
            jobs_written = write_5030_section_data(worksheet, material_data, start_row)
            summary['jobs_updated']['material'] = jobs_written
        
        # Save the workbook with proper error handling
        try:
            workbook.save(file_path)
            logging.info(f"Successfully saved workbook to {file_path}")
            
            # Verify the saved file by trying to reopen it
            try:
                test_wb = load_workbook(file_path, keep_vba=True)
                test_wb.close()
                logging.info(f"Verified saved file can be reopened successfully")
                summary['success'] = True
            except Exception as verify_error:
                raise Exception(f"Saved file appears to be corrupted - verification failed: {str(verify_error)}")
                
        except Exception as save_error:
            # Check if it's a merged cell error and provide specific guidance
            if "MergedCell" in str(save_error):
                raise Exception(f"Failed to save workbook due to merged cell conflict. Please ensure no merged cells are being written to. Error: {str(save_error)}")
            else:
                raise Exception(f"Failed to save workbook: {str(save_error)}")
        
        logging.info(f"Successfully updated WIP Report: {summary}")
        
    except Exception as e:
        error_msg = f"Error updating WIP Report: {str(e)}"
        logging.error(error_msg)
        summary['error'] = error_msg
        
        # If we created a backup and there was an error, we could restore it here
        if summary['backup_created'] and create_backup_flag:
            try:
                shutil.copy2(summary['backup_created'], file_path)
                logging.info(f"Restored original file from backup due to error")
            except Exception as restore_error:
                logging.error(f"Failed to restore backup: {str(restore_error)}")
    
    return summary


if __name__ == "__main__":
    # Test with sample data
    logging.basicConfig(level=logging.INFO)
    print("Excel Integration v2 module loaded successfully") 