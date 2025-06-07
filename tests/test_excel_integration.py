"""
Test cases for Excel Integration Module

This module contains pytest test cases to validate the Excel integration functionality.
"""

import pytest
import pandas as pd
import tempfile
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.cell.cell import Cell

from src.data_processing.excel_integration import (
    load_wip_workbook,
    find_or_create_monthly_tab,
    find_section_markers,
    detect_data_region,
    is_formula_cell,
    clear_data_preserve_formulas,
    write_job_data_to_section,
    create_backup,
    get_existing_data_from_section
)


@pytest.fixture
def sample_workbook():
    """Create a sample workbook for testing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Jan 24"
    
    # Add some sample data with section markers
    ws['A1'] = "Job Number"
    ws['B1'] = "Material Cost"
    ws['A3'] = "5040"  # Material section marker
    ws['A4'] = "JOB001"
    ws['B4'] = 10000
    ws['A5'] = "JOB002"
    ws['B5'] = 5000
    
    ws['A8'] = "5030"  # Labor section marker
    ws['A9'] = "JOB001"
    ws['B9'] = 8000
    ws['A10'] = "JOB002"
    ws['B10'] = 4000
    
    # Add a formula cell
    ws['C4'] = "=B4*1.1"
    
    return wb


@pytest.fixture
def sample_excel_file(sample_workbook):
    """Create a temporary Excel file for testing."""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
        sample_workbook.save(tmp_file.name)
        yield tmp_file.name
    
    # Cleanup
    os.unlink(tmp_file.name)


class TestLoadWIPWorkbook:
    """Test cases for load_wip_workbook function."""
    
    def test_load_workbook_success(self, sample_excel_file):
        """Test successful loading of workbook."""
        workbook = load_wip_workbook(sample_excel_file)
        
        assert workbook is not None
        assert len(workbook.sheetnames) > 0
        assert "Jan 24" in workbook.sheetnames
    
    def test_load_workbook_file_not_found(self):
        """Test error handling when file doesn't exist."""
        with pytest.raises(FileNotFoundError):
            load_wip_workbook("nonexistent_file.xlsx")
    
    def test_load_workbook_keep_vba_flag(self, sample_excel_file):
        """Test that keep_vba flag is respected."""
        # This test mainly ensures the parameter is passed correctly
        workbook = load_wip_workbook(sample_excel_file, keep_vba=False)
        assert workbook is not None


class TestFindOrCreateMonthlyTab:
    """Test cases for find_or_create_monthly_tab function."""
    
    def test_find_existing_tab(self, sample_workbook):
        """Test finding an existing monthly tab."""
        worksheet = find_or_create_monthly_tab(sample_workbook, "Jan 24")
        
        assert worksheet is not None
        assert worksheet.title == "Jan 24"
    
    def test_create_new_tab(self, sample_workbook):
        """Test creating a new monthly tab."""
        worksheet = find_or_create_monthly_tab(sample_workbook, "Feb 24")
        
        assert worksheet is not None
        assert worksheet.title == "Feb 24"
        assert "Feb 24" in sample_workbook.sheetnames
    
    def test_create_tab_with_template(self):
        """Test creating a new tab when template exists."""
        wb = Workbook()
        template_sheet = wb.active
        template_sheet.title = "Template"
        template_sheet['A1'] = "Template Data"
        
        new_worksheet = find_or_create_monthly_tab(wb, "Mar 24")
        
        assert new_worksheet.title == "Mar 24"
        assert new_worksheet['A1'].value == "Template Data"


class TestFindSectionMarkers:
    """Test cases for find_section_markers function."""
    
    def test_find_existing_markers(self, sample_workbook):
        """Test finding existing section markers."""
        worksheet = sample_workbook.active
        markers = find_section_markers(worksheet, ['5040', '5030'])
        
        assert '5040' in markers
        assert '5030' in markers
        assert markers['5040'] == (3, 1)  # Row 3, Column 1
        assert markers['5030'] == (8, 1)  # Row 8, Column 1
    
    def test_find_nonexistent_markers(self, sample_workbook):
        """Test behavior when markers don't exist."""
        worksheet = sample_workbook.active
        markers = find_section_markers(worksheet, ['9999', '8888'])
        
        assert markers['9999'] is None
        assert markers['8888'] is None
    
    def test_find_mixed_markers(self, sample_workbook):
        """Test finding mix of existing and non-existing markers."""
        worksheet = sample_workbook.active
        markers = find_section_markers(worksheet, ['5040', '9999'])
        
        assert markers['5040'] == (3, 1)
        assert markers['9999'] is None


class TestDetectDataRegion:
    """Test cases for detect_data_region function."""
    
    def test_detect_data_region_basic(self, sample_workbook):
        """Test basic data region detection."""
        worksheet = sample_workbook.active
        end_row, end_col = detect_data_region(worksheet, 4, 1)  # Start after 5040 marker
        
        # Should detect data in rows 4-5, columns 1+
        assert end_row >= 5
        assert end_col >= 1
    
    def test_detect_data_region_empty_area(self):
        """Test data region detection in empty area."""
        wb = Workbook()
        ws = wb.active
        
        end_row, end_col = detect_data_region(ws, 1, 1)
        
        # Should return starting position when no data found
        assert end_row == 1
        assert end_col == 1
    
    def test_detect_data_region_with_gaps(self):
        """Test data region detection with gaps in data."""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Data1"
        ws['A2'] = "Data2"
        # A3 is empty
        ws['A4'] = "Data4"
        
        end_row, end_col = detect_data_region(ws, 1, 1, max_rows=10)
        
        # Should detect up to row 4 despite the gap
        assert end_row >= 2  # At least to row 2


class TestIsFormulaCell:
    """Test cases for is_formula_cell function."""
    
    def test_formula_cell_with_equals(self):
        """Test detection of formula cell starting with =."""
        wb = Workbook()
        ws = wb.active
        cell = ws['A1']
        cell.value = "=SUM(B1:B10)"
        
        assert is_formula_cell(cell) is True
    
    def test_formula_cell_data_type(self):
        """Test detection based on data type."""
        wb = Workbook()
        ws = wb.active
        cell = ws['A1']
        cell.value = "=B1*2"
        # Simulate formula data type
        cell.data_type = 'f'
        
        assert is_formula_cell(cell) is True
    
    def test_non_formula_cell(self):
        """Test detection of non-formula cell."""
        wb = Workbook()
        ws = wb.active
        cell = ws['A1']
        cell.value = "Regular Text"
        
        assert is_formula_cell(cell) is False
    
    def test_numeric_cell(self):
        """Test detection of numeric cell."""
        wb = Workbook()
        ws = wb.active
        cell = ws['A1']
        cell.value = 12345
        
        assert is_formula_cell(cell) is False


class TestClearDataPreserveFormulas:
    """Test cases for clear_data_preserve_formulas function."""
    
    def test_clear_data_preserve_formulas(self):
        """Test clearing data while preserving formulas."""
        wb = Workbook()
        ws = wb.active
        
        # Set up test data
        ws['A1'] = "Text Data"
        ws['A2'] = 12345
        ws['A3'] = "=SUM(B1:B10)"  # Formula
        ws['A4'] = "More Text"
        
        # Clear range that includes formula
        cells_cleared = clear_data_preserve_formulas(ws, 1, 4, 1, 1)
        
        # Should clear non-formula cells but preserve formula
        assert ws['A1'].value is None  # Cleared
        assert ws['A2'].value is None  # Cleared
        assert ws['A3'].value == "=SUM(B1:B10)"  # Preserved
        assert ws['A4'].value is None  # Cleared
        assert cells_cleared == 3  # 3 non-formula cells cleared
    
    def test_clear_empty_range(self):
        """Test clearing empty range."""
        wb = Workbook()
        ws = wb.active
        
        cells_cleared = clear_data_preserve_formulas(ws, 1, 3, 1, 3)
        
        assert cells_cleared == 0


class TestWriteJobDataToSection:
    """Test cases for write_job_data_to_section function."""
    
    def test_write_material_data(self):
        """Test writing material data to section."""
        wb = Workbook()
        ws = wb.active
        
        job_data = pd.DataFrame({
            'Job Number': ['JOB001', 'JOB002'],
            'Material': [10000, 5000]
        })
        
        jobs_written = write_job_data_to_section(ws, job_data, 1, 1)
        
        assert jobs_written == 2
        assert ws['A2'].value == 'JOB001'
        assert ws['B2'].value == 10000
        assert ws['A3'].value == 'JOB002'
        assert ws['B3'].value == 5000
    
    def test_write_labor_data(self):
        """Test writing labor data to section."""
        wb = Workbook()
        ws = wb.active
        
        job_data = pd.DataFrame({
            'Job Number': ['JOB001', 'JOB002'],
            'Labor': [8000, 4000]
        })
        
        jobs_written = write_job_data_to_section(ws, job_data, 1, 1)
        
        assert jobs_written == 2
        assert ws['A2'].value == 'JOB001'
        assert ws['B2'].value == 8000
    
    def test_write_data_skip_formula_cells(self):
        """Test that writing skips formula cells."""
        wb = Workbook()
        ws = wb.active
        
        # Set up formula cells
        ws['A2'] = "=A1"
        ws['B2'] = "=B1"
        
        job_data = pd.DataFrame({
            'Job Number': ['JOB001'],
            'Material': [10000]
        })
        
        jobs_written = write_job_data_to_section(ws, job_data, 1, 1)
        
        # Should not overwrite formula cells
        assert ws['A2'].value == "=A1"
        assert ws['B2'].value == "=B1"
        assert jobs_written == 0  # No jobs written due to formula cells
    
    def test_write_data_no_recognized_column(self):
        """Test behavior when no recognized data column exists."""
        wb = Workbook()
        ws = wb.active
        
        job_data = pd.DataFrame({
            'Job Number': ['JOB001'],
            'Unknown Column': [10000]
        })
        
        jobs_written = write_job_data_to_section(ws, job_data, 1, 1)
        
        assert jobs_written == 0


class TestCreateBackup:
    """Test cases for create_backup function."""
    
    def test_create_backup_success(self, sample_excel_file):
        """Test successful backup creation."""
        with tempfile.TemporaryDirectory() as temp_dir:
            backup_path = create_backup(sample_excel_file, temp_dir)
            
            assert os.path.exists(backup_path)
            assert "BACKUP" in backup_path
            assert backup_path.endswith('.xlsx')
    
    def test_create_backup_creates_directory(self, sample_excel_file):
        """Test that backup creates directory if it doesn't exist."""
        with tempfile.TemporaryDirectory() as temp_dir:
            backup_dir = os.path.join(temp_dir, "new_backup_dir")
            backup_path = create_backup(sample_excel_file, backup_dir)
            
            assert os.path.exists(backup_dir)
            assert os.path.exists(backup_path)
    
    def test_create_backup_nonexistent_file(self):
        """Test backup creation with nonexistent source file."""
        with pytest.raises(Exception):
            create_backup("nonexistent_file.xlsx")


class TestGetExistingDataFromSection:
    """Test cases for get_existing_data_from_section function."""
    
    def test_get_existing_data_success(self, sample_workbook):
        """Test successful extraction of existing data."""
        worksheet = sample_workbook.active
        df = get_existing_data_from_section(worksheet, '5040')
        
        # Should have data rows (excluding section headers)
        assert len(df) >= 2
        assert 'Job Number' in df.columns
        assert 'Current Value' in df.columns
        assert 'JOB001' in df['Job Number'].values
        assert 'JOB002' in df['Job Number'].values
    
    def test_get_existing_data_nonexistent_section(self, sample_workbook):
        """Test extraction from nonexistent section."""
        worksheet = sample_workbook.active
        df = get_existing_data_from_section(worksheet, '9999')
        
        assert len(df) == 0
    
    def test_get_existing_data_empty_section(self):
        """Test extraction from empty section."""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '5040'  # Section marker but no data
        
        df = get_existing_data_from_section(ws, '5040')
        
        assert len(df) == 0


if __name__ == "__main__":
    # Run tests if executed directly
    pytest.main([__file__]) 